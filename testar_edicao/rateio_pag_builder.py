import re
import unicodedata
from pathlib import Path
from typing import Union, Tuple
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import numbers

# ==========================================
#              FUNÇÕES ÚTEIS
# ==========================================

def _strip_accents(s: str) -> str:
    if pd.isna(s) or s is None:
        return ''
    s = unicodedata.normalize('NFKD', str(s))
    return ''.join(ch for ch in s if not unicodedata.combining(ch))

def _norm_colname(s: str) -> str:
    return _strip_accents(str(s)).lower().strip()

def _clean_str(x) -> str:
    if pd.isna(x): return ''
    if isinstance(x, float):
        if x.is_integer(): return str(int(x))
        return str(x)
    s = str(x).strip()
    if s.endswith('.0'): s = s[:-2]
    return s

def _norm_coletor(x: str) -> str:
    up = _strip_accents(_clean_str(x)).upper()
    return re.sub(r'[^A-Z0-9]', '', up)

def _is_valid_coletor(coletor: str) -> bool:
    if pd.isna(coletor) or str(coletor).strip() == '' or str(coletor).strip().lower() == 'nan':
        return False
    c = _strip_accents(_clean_str(coletor)).upper()
    c_clean = re.sub(r'[^A-Z0-9]', '', c)
    if re.fullmatch(r'\d{6,}', c_clean): 
        return True
    if 6 <= len(c_clean) <= 12 and any(x.isalpha() for x in c_clean) and any(x.isdigit() for x in c_clean): 
        return True
    return False

def _tipo_de_coletor(coletor: str) -> str:
    c = _norm_coletor(coletor)
    if c == 'SEMCENTRODECUSTO': return '-'
    if re.fullmatch(r'\d+', c): return 'N'
    if any(x.isalpha() for x in c) and any(x.isdigit() for x in c): return 'K'
    return ''

def _clean_valor_series(s: pd.Series) -> pd.Series:
    def limpa_valor(val):
        if pd.isna(val) or str(val).strip() == '': return None
        if isinstance(val, (int, float)): return float(val)
        v = str(val).upper().replace('R$', '').replace('\xa0', '').replace(' ', '').strip()
        if '.' in v and ',' in v: v = v.replace('.', '').replace(',', '.')
        elif ',' in v: v = v.replace(',', '.')
        try: return float(v)
        except ValueError: return None
    return s.apply(limpa_valor)

# ==========================================
#     LEITURA: RATEIO RECEBIDO (ISOLAMENTO)
# ==========================================

def ler_rr_bruto(caminho_rr: Union[str, Path]) -> pd.DataFrame:
    xls = pd.ExcelFile(caminho_rr, engine='openpyxl')
    frames = []
    
    for sh in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name=sh, header=None, engine='openpyxl')
        idx_ancora = -1
        max_knf = 0
        
        for j in range(df.shape[1]):
            col_str = df.iloc[:, j].astype(str).str.strip().str.upper()
            count_knf = col_str.isin(['K', 'N', 'F']).sum()
            if count_knf > max_knf and count_knf >= 2:
                max_knf = count_knf
                idx_ancora = j
                
        if idx_ancora == -1: continue
            
        df_regional = df.iloc[:, idx_ancora:].copy()
        df_regional.columns = range(df_regional.shape[1])
        
        col_coletor, col_subnum, col_valor = -1, -1, -1
        header_row_idx = -1
        
        for i, row in df_regional.head(20).iterrows():
            row_norm = [_norm_colname(str(x)) for x in row.values]
            if 'coletor' in row_norm or 'subnumero' in row_norm:
                for j in range(len(row_norm)):
                    val = row_norm[j]
                    if val == 'coletor' and col_coletor == -1: col_coletor = j
                    elif 'subnumero' in val and col_subnum == -1: col_subnum = j
                    elif 'valor' in val and 'servico' not in val and col_valor == -1: col_valor = j
                
                if col_valor == -1:
                    for j in range(len(row_norm)):
                        if 'valor' in row_norm[j] and col_valor == -1: col_valor = j
                        
                if col_coletor != -1 and col_valor != -1:
                    header_row_idx = i
                    break

        if header_row_idx != -1:
            tmp = df_regional.iloc[header_row_idx + 1:].copy()
            def safe_get(c): return tmp.iloc[:, c] if c != -1 and c < tmp.shape[1] else pd.Series(['']*len(tmp))
            tmp_clean = pd.DataFrame({
                'COLETOR_ORIG': safe_get(col_coletor),
                'SUBNUM_ORIG': safe_get(col_subnum),
                'VALOR': safe_get(col_valor)
            })
        else:
            col_coletor_tb = -1
            max_validos = 0
            for j in range(1, df_regional.shape[1]):
                mask = df_regional.iloc[:, j].astype(str).apply(_is_valid_coletor)
                qtd = mask.sum()
                if qtd > max_validos and qtd >= 2:
                    max_validos = qtd
                    col_coletor_tb = j
                    
            col_valor_tb = -1
            max_nums = 0
            for j in range(1, df_regional.shape[1]):
                if j == col_coletor_tb: continue
                nums = _clean_valor_series(df_regional.iloc[:, j]).notna().sum()
                if nums > max_nums:
                    max_nums = nums
                    col_valor_tb = j
                    
            if col_coletor_tb == -1 or col_valor_tb == -1: continue
                
            tmp_clean = pd.DataFrame({
                'COLETOR_ORIG': df_regional.iloc[:, col_coletor_tb],
                'SUBNUM_ORIG': pd.Series(['']*len(df_regional)), 
                'VALOR': df_regional.iloc[:, col_valor_tb]
            })

        tmp_clean['VALOR'] = _clean_valor_series(tmp_clean['VALOR'])
        tmp_clean = tmp_clean.dropna(subset=['VALOR'])
        
        def get_best_coletor(r):
            sub = _clean_str(r['SUBNUM_ORIG'])
            if _is_valid_coletor(sub): return sub
            col = _clean_str(r['COLETOR_ORIG'])
            if _is_valid_coletor(col): return col
            return None
            
        tmp_clean['COLETOR_FINAL'] = tmp_clean.apply(get_best_coletor, axis=1)
        tmp_clean = tmp_clean.dropna(subset=['COLETOR_FINAL'])
        
        if not tmp_clean.empty:
            tmp_clean['COLETOR'] = tmp_clean['COLETOR_FINAL'].apply(_norm_coletor)
            tmp_clean['TIPOCOLETOR'] = tmp_clean['COLETOR'].apply(_tipo_de_coletor)
            frames.append(tmp_clean[['TIPOCOLETOR', 'COLETOR', 'VALOR']])

    if frames:
        return pd.concat(frames, ignore_index=True)
    return pd.DataFrame(columns=['TIPOCOLETOR', 'COLETOR', 'VALOR'])

# ==========================================
#     LEITURA: CORREIOS (COLA-PEDAÇOS)
# ==========================================

def _extrair_coletor_de_titular(texto: str) -> str:
    if pd.isna(texto) or str(texto).strip() == '': return "SEM CENTRO DE CUSTO"
    t = _strip_accents(str(texto)).upper()
    
    m = re.search(r'(?<!\d)(\d{6,})(?!\d)', t)
    if m: return _norm_coletor(m.group(1))
    
    palavras = t.split()
    for p in palavras:
        p_clean = re.sub(r'[^A-Z0-9]', '', p)
        if 6 <= len(p_clean) <= 12 and any(c.isalpha() for c in p_clean) and any(c.isdigit() for c in p_clean):
            return p_clean
            
    for i in range(len(palavras) - 1):
        p1 = re.sub(r'[^A-Z0-9]', '', palavras[i])
        p2 = re.sub(r'[^A-Z0-9]', '', palavras[i+1])
        comb = p1 + p2
        if 8 <= len(comb) <= 12 and any(c.isalpha() for c in comb) and any(c.isdigit() for c in comb):
            return comb
            
    return "SEM CENTRO DE CUSTO"

def ler_correios_bruto(caminho_correios: Union[str, Path]) -> Tuple[pd.DataFrame, float]:
    df_raw = pd.read_excel(caminho_correios, header=None, engine='openpyxl')
    idx_header = -1
    col_titular, col_valor = -1, -1
    
    for i, row in df_raw.head(20).iterrows():
        row_norm = [_norm_colname(str(x)) for x in row.values]
        if any('titular do cartao' in c for c in row_norm) and any('valor do servico' in c for c in row_norm):
            idx_header = i
            for j, c in enumerate(row_norm):
                if 'titular do cartao' in c: col_titular = j
                if 'valor do servico' in c: col_valor = j
            break
            
    valor_liquido = 0.0
    idx_fim_tabela = len(df_raw)
    
    for i in range(len(df_raw) - 1, -1, -1):
        row_norm = [_norm_colname(str(x)) for x in df_raw.iloc[i].values]
        if any('valor liquido' in c for c in row_norm):
            idx_fim_tabela = i
            for j, c in enumerate(row_norm):
                if 'valor liquido' in c:
                    if i + 1 < len(df_raw):
                        val_raw = df_raw.iloc[i + 1, j]
                        valor_liquido = _clean_valor_series(pd.Series([val_raw])).iloc[0]
                    break
            break

    if idx_header == -1: 
        return pd.DataFrame(columns=['TIPOCOLETOR', 'COLETOR', 'VALOR']), valor_liquido
    
    df = df_raw.iloc[idx_header + 1 : idx_fim_tabela, [col_titular, col_valor]].copy()
    df.columns = ['TITULAR', 'VALOR']
    
    # IGNORA LINHAS DE ENCARGOS E DESCONTOS (Serão tratados matematicamente no final)
    mask_ignorar = df['TITULAR'].astype(str).str.upper().str.contains('ENCARGO|DESCONTO|CREDITO')
    df = df[~mask_ignorar].copy()
    
    df['VALOR'] = _clean_valor_series(df['VALOR'])
    df = df.dropna(subset=['VALOR'])
    df['COLETOR'] = df['TITULAR'].apply(_extrair_coletor_de_titular)
    df['TIPOCOLETOR'] = df['COLETOR'].apply(_tipo_de_coletor)
    
    return df[['TIPOCOLETOR', 'COLETOR', 'VALOR']], valor_liquido

# ==========================================
#     CONSTRUÇÃO E CONCILIAÇÃO FINAL
# ==========================================

def _formatar_planilha_final(arquivo_xlsx: Union[str, Path], sheet='Planilha1'):
    wb = load_workbook(arquivo_xlsx)
    if sheet not in wb.sheetnames:
        wb.close(); return
    ws = wb[sheet]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        idx_valor = header.index('VALOR') + 1
        idx_op = header.index('OPERACAO') + 1
    except ValueError:
        wb.close(); return
        
    for row in ws.iter_rows(min_row=2):
        cell_valor = row[idx_valor - 1]
        if isinstance(cell_valor.value, (int, float)):
            cell_valor.number_format = numbers.FORMAT_NUMBER_00
            
    wb.save(arquivo_xlsx)
    wb.close()

def gerar_rateio_pag(
    caminho_correios: Union[str, Path],
    caminho_rr: Union[str, Path],
    saida: Union[str, Path] = 'RATEIO PAG.xlsx',
    operacao_para_diagrama: int = 10,
    tolerancia_igual: float = 0.05, 
    debug: bool = True
) -> pd.DataFrame:

    df_rr_raw = ler_rr_bruto(caminho_rr)
    df_corr_raw, valor_liquido_correios = ler_correios_bruto(caminho_correios)

    # GARANTIA DE TIPO NUMÉRICO
    df_rr_raw['VALOR'] = pd.to_numeric(df_rr_raw['VALOR'], errors='coerce').fillna(0.0)
    df_corr_raw['VALOR'] = pd.to_numeric(df_corr_raw['VALOR'], errors='coerce').fillna(0.0)

    total_rr = float(df_rr_raw['VALOR'].sum()) if not df_rr_raw.empty else 0.0
    total_corr_soma = float(df_corr_raw['VALOR'].sum()) if not df_corr_raw.empty else 0.0
    total_corr = valor_liquido_correios if valor_liquido_correios > 0 else total_corr_soma

    if debug:
        print(f"[DEBUG] TOTAL RR               = R$ {total_rr:.2f}")
        print(f"[DEBUG] TOTAL CORREIOS (SOMA)  = R$ {total_corr_soma:.2f}")
        print(f"[DEBUG] TOTAL CORREIOS (LÍQ)   = R$ {valor_liquido_correios:.2f}")
        print(f"[DEBUG] DIFERENÇA (LÍQ - RR)   = R$ {round(total_corr - total_rr, 2):.2f}")

    # 1. BASE INICIAL: Tudo que a regional aprovou
    linhas_finais = df_rr_raw.to_dict('records') if not df_rr_raw.empty else []

    if abs(total_corr - total_rr) > tolerancia_igual:
        if debug: print("[DEBUG] Valores divergem. Buscando pacotes exatos faltantes...")
        
        rr_valores_exatos = df_rr_raw['VALOR'].round(2).tolist()
        
        # Saldo por Centro de Custo
        saldo_rr_cc = {}
        for _, row in df_rr_raw.iterrows():
            c = row['COLETOR']
            saldo_rr_cc[c] = saldo_rr_cc.get(c, 0.0) + row['VALOR']
            
        # A "Carteira" Global (Garante que nunca vamos adicionar pacotes a mais)
        saldo_global_rr = total_rr
        pacotes_faltantes = []
        
        for _, row in df_corr_raw.iterrows():
            c = row['COLETOR']
            v = round(row['VALOR'], 2)
            tipo = row['TIPOCOLETOR']
            
            matched = False
            
            # Fase 1: Tenta pagar usando o saldo do próprio Centro de Custo
            if c in saldo_rr_cc and saldo_rr_cc[c] >= v - 0.02:
                saldo_rr_cc[c] -= v
                saldo_global_rr -= v
                matched = True
                # Remove da lista de valores exatos para não usar duas vezes
                for i, rr_v in enumerate(rr_valores_exatos):
                    if abs(rr_v - v) <= 0.02:
                        rr_valores_exatos.pop(i)
                        break
            
            # Fase 2: Tenta pagar achando o valor exato (caso o CC tenha mudado de nome)
            if not matched:
                for i, rr_v in enumerate(rr_valores_exatos):
                    if abs(rr_v - v) <= 0.02:
                        rr_valores_exatos.pop(i)
                        saldo_global_rr -= v
                        matched = True
                        break
                        
            # Se não achou de jeito nenhum, vai para a fila de espera
            if not matched:
                pacotes_faltantes.append({
                    'TIPOCOLETOR': tipo,
                    'COLETOR': c,
                    'VALOR': v
                })
                
        # Fase 3: Absorção pela Carteira Global
        # Se a regional agrupou valores E mudou o nome, o pacote cai aqui.
        # Usamos o dinheiro que sobrou na carteira para pagar. Se a carteira zerar, aí sim é um pacote faltante!
        pacotes_adicionados = 0
        for pct in pacotes_faltantes:
            v = pct['VALOR']
            if saldo_global_rr >= v - 0.02:
                # A carteira cobre! (Falso positivo, a regional já pagou por isso)
                saldo_global_rr -= v
            else:
                # A carteira não cobre! Realmente faltou na planilha da regional.
                linhas_finais.append(pct)
                pacotes_adicionados += 1
                
        if debug: print(f"[DEBUG] Adicionados {pacotes_adicionados} pacotes exatos dos Correios.")

    # Transforma em DataFrame e agrupa para unificar linhas do mesmo CC
    final_base = pd.DataFrame(linhas_finais)
    if not final_base.empty:
        final_base = final_base.groupby(['TIPOCOLETOR', 'COLETOR'], as_index=False)['VALOR'].sum()

    # 3. RATEIO DE ENCARGOS E DESCONTOS (Matemática)
    soma_atual = final_base['VALOR'].sum() if not final_base.empty else 0.0
    diferenca_rateio = round(total_corr - soma_atual, 2)
    
    # Se sobrou diferença (Encargos) ou faltou (Descontos), rateia entre todos
    if abs(diferenca_rateio) > 0.02 and not final_base.empty:
        if debug: print(f"[DEBUG] Rateando R$ {diferenca_rateio:.2f} (Encargos/Descontos) proporcionalmente...")
        
        soma_validos = final_base['VALOR'].sum()
        if soma_validos > 0:
            # Calcula a proporção
            final_base['VALOR_ADD'] = (final_base['VALOR'] / soma_validos) * diferenca_rateio
            final_base['VALOR_ADD'] = final_base['VALOR_ADD'].round(2)
            
            # Ajuste de centavos (dízimas)
            diff_centavos = round(diferenca_rateio - final_base['VALOR_ADD'].sum(), 2)
            if diff_centavos != 0:
                idx_max = final_base['VALOR'].idxmax() 
                final_base.loc[idx_max, 'VALOR_ADD'] += diff_centavos 
                
            # Aplica o rateio
            final_base['VALOR'] += final_base['VALOR_ADD']
            final_base = final_base.drop(columns=['VALOR_ADD'])

    # Monta Tabela Final
    final = pd.DataFrame()
    if not final_base.empty:
        final['ITEM'] = [1] * len(final_base)
        final['TIPOCOLETOR'] = final_base['TIPOCOLETOR']
        final['COLETOR'] = final_base['COLETOR']
        final['OPERACAO'] = final_base['TIPOCOLETOR'].apply(lambda t: operacao_para_diagrama if t == 'N' else '')
        final['SUBNUMERO'] = ''
        final['VALOR'] = final_base['VALOR']
        final['DESCRICAO'] = ''

        # Ordenação (K -> N -> Faltantes no Final)
        final['__ord'] = final['TIPOCOLETOR'].map({'K': 0, 'N': 1}).fillna(2)
        final = final.sort_values(['__ord', 'COLETOR']).drop(columns='__ord').reset_index(drop=True)

    saida = Path(saida)
    with pd.ExcelWriter(saida, engine='openpyxl') as writer:
        final.to_excel(writer, sheet_name='Planilha1', index=False)

    _formatar_planilha_final(saida, 'Planilha1')

    if debug: print(f"[DEBUG] Arquivo gerado com sucesso: {saida.resolve()}")

    return final

# ==========================================
#     INICIALIZAÇÃO AUTOMÁTICA
# ==========================================
if __name__ == '__main__':
    pasta_trabalho = Path(r"C:/Users/pedro.henrsilva/OneDrive - MRV/Área de Trabalho/Rateio_Regionais/testar_edicao")
    
    caminho_rr = None
    caminho_correios = None
    
    for ficheiro in pasta_trabalho.glob('*.xlsx'):
        nome_ficheiro = ficheiro.name.upper()
        if nome_ficheiro.startswith('~$') or nome_ficheiro == 'RATEIO PAG.XLSX': continue
        if 'RATEIO RECEBIDO' in nome_ficheiro: caminho_rr = ficheiro
        elif re.match(r'^\d{7}\.XLSX$', nome_ficheiro): caminho_correios = ficheiro

    if not caminho_rr or not caminho_correios:
        print("ERRO: Faltam planilhas na pasta!")
    else:
        print(f"\n>> Iniciando processamento...")
        caminho_saida = pasta_trabalho / "RATEIO PAG.xlsx"
        final = gerar_rateio_pag(caminho_correios=caminho_correios, caminho_rr=caminho_rr, saida=caminho_saida)
        print(f"\nTotal de linhas geradas no final: {len(final)}")
