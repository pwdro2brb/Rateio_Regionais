import tkinter as tk
from tkinter import ttk, messagebox
import threading
import sys
import time

# ==============================================================================
# 1. IMPORTAÇÕES DOS SEUS CÓDIGOS
# ==============================================================================
# Cole aqui todos os seus imports (pandas, selenium, openpyxl, win32com, etc.)
import os
import re
import time
import getpass
import unicodedata
import pandas as pd
from pathlib import Path
from PyPDF2 import PdfReader
from typing import Optional, Dict, List

# --- Selenium Imports ---
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import (
    StaleElementReferenceException, TimeoutException, 
    ElementClickInterceptedException, ElementNotInteractableException, WebDriverException
)
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys

# --- Pywinauto Imports ---
from pywinauto.application import Application
from pywinauto import Desktop

#funções de ajuda para Selenium e Windows (reutilizáveis e específicas para os desafios deste site)

import time
from selenium.webdriver.common.by import By
from selenium.common.exceptions import TimeoutException
from datetime import datetime
from pywinauto import Desktop
from pywinauto.timings import wait_until
import traceback

#-------------------------------------------------------------------------------------------------

import win32com.client as win32
from datetime import datetime, timedelta
#-------------------------------------------------------------------------------------------------

from typing import Union, Tuple
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import numbers

CNPJ_CORREIOS_FIXO = "34028316001509"   
DATE_RE = r"([0-3]?\d/[01]?\d/\d{4})"   

# ==============================================================================
# 2. SUAS FUNÇÕES PRINCIPAIS (ETAPAS 1, 2 e 3)
# ==============================================================================

def criar_rascunhos_correios():
    print("Iniciando a criação de rascunhos no Outlook...")
    caminho_base = r"\\Bhz-fls-app1\mrvbh\Gerência Administrativa\Pública\NUCLEO DE CONTRATOS E APOIO A GESTÃO\CONTRATOS\Contratos Serviços\1. CORREIOS\2. Faturamento\2026"
    
    pastas_meses = [f for f in os.listdir(caminho_base) if os.path.isdir(os.path.join(caminho_base, f))]
    if not pastas_meses:
        print("Nenhuma pasta de mês encontrada no diretório.")
        return

    pastas_meses.sort()
    pasta_mes_recente = pastas_meses[-1]
    caminho_mes_recente = os.path.join(caminho_base, pasta_mes_recente)
    nome_mes = pasta_mes_recente.split("-")[-1].strip()

    print(f"Pasta mais recente encontrada: {pasta_mes_recente}")

    contatos_para = {
        "Campinas": "flavia.pinho@mrv.com.br; ana.tilli@mrv.com.br",
        "Ribeirão Preto": "kaylana.alves@mrv.com.br",
        "Centro Oeste": "nicole.souza@mrv.com.br; maksuel.araujo@mrv.com.br; eunice.prudente@primeconstrucoes.com.br; maryanne.camargo@primeconstrucoes.com.br",
        "Nordeste": "langela.santos@mrv.com.br",
        "Sul": "victoria.gomes@mrv.com.br; filipe.avila@mrv.com.br; simone.csantos@mrv.com.br; monique.silva@mrv.com.br",
        "São Paulo": "telma.amattos@mrv.com.br; cristina.demetrio@parceiro.mrv.com.br; manoella.camargo@mrv.com.br; luciano.lsilva@mrv.com.br; nicoli.santos@mrv.com.br",
        "Triângulo": "kamilly.silva@mrv.com.br; kaylana.alves@mrv.com.br; maria.fernnanda@mrv.com.br"
    }

    agora = datetime.now()
    saudacao = "Bom dia" if agora.hour < 12 else "Boa tarde"
    prazo_rateio = agora + timedelta(hours=32)
    prazo_formatado = prazo_rateio.strftime("%d/%m/%Y às %H:%M")

    corpo_email = f"""
    <p style="font-family: Calibri, Arial, sans-serif; font-size: 11pt; color: #000000;">
        {saudacao}, Prezado(s)!<br><br>
        Segue em anexo o extrato dos Correios. O rateio deverá ser enviado até <b>{prazo_formatado}</b>.<br><br>
        Atenciosamente,
    </p>
    """

    outlook = win32.Dispatch('outlook.application')
    pastas_regionais = os.listdir(caminho_mes_recente)
    
    for regional in pastas_regionais:
        caminho_regional = os.path.join(caminho_mes_recente, regional)
        if not os.path.isdir(caminho_regional) or regional.upper() == "BH":
            continue

        print(f"Gerando rascunho para: {regional}...")
        mail = outlook.CreateItem(0)
        mail.To = contatos_para.get(regional, "")
        
        cc_padrao = "vanessa.brodrigues@mrv.com.br; correiosbh@mrv.com.br"
        if regional in ["Triângulo", "Ribeirão Preto"]:
            mail.CC = f"{cc_padrao}; conceicao@mrv.com.br"
        else:
            mail.CC = cc_padrao

        mail.Subject = f"RES: Extrato Correios - {regional} ({nome_mes})"

        arquivos_na_pasta = os.listdir(caminho_regional)
        for arquivo in arquivos_na_pasta:
            caminho_arquivo = os.path.join(caminho_regional, arquivo)
            if os.path.isfile(caminho_arquivo):
                mail.Attachments.Add(caminho_arquivo)

        mail.Display() 
        assinatura_outlook = mail.HTMLBody
        mail.HTMLBody = f"<html><body>{corpo_email}{assinatura_outlook}</body></html>"
        mail.Save()
        mail.Close(0)

    print("\nProcesso concluído! Verifique a pasta 'Rascunhos' no seu Outlook.")


def preparar_e_gerar_rateio():
    print("Lendo planilhas na pasta testar_edicao...")
    pasta_trabalho = Path(r"C:\Users\pedro.henrsilva\OneDrive - MRV\Área de Trabalho\Rateio_Regionais\testar_edicao")
    
    caminho_rr = None
    caminho_correios = None
    
    for ficheiro in pasta_trabalho.glob('*.xlsx'):
        nome_ficheiro = ficheiro.name.upper()
        if nome_ficheiro.startswith('~$') or nome_ficheiro == 'RATEIO PAG.XLSX': continue
        if 'RATEIO RECEBIDO' in nome_ficheiro: caminho_rr = ficheiro
        elif re.match(r'^\d{7}\.XLSX$', nome_ficheiro): caminho_correios = ficheiro

    if not caminho_rr or not caminho_correios:
        print("ERRO: Faltam planilhas na pasta testar_edicao!")
        return

    print(f">> Iniciando processamento...")
    caminho_saida = pasta_trabalho / "RATEIO PAG.xlsx"
    
    # Chama a sua função gerar_rateio_pag que está mais abaixo no código
    final = gerar_rateio_pag(caminho_correios=caminho_correios, caminho_rr=caminho_rr, saida=caminho_saida)
    
    print(f"Total de linhas geradas no final: {len(final)}")
    print("Arquivo RATEIO PAG.xlsx gerado com sucesso!")


def lancar_nota_fiscal():
    print("Iniciando robô de lançamento...")
    
    # 1. BUSCADOR DE ARQUIVOS
    PASTA_BASE = Path(r"C:\Users\pedro.henrsilva\OneDrive - MRV\Área de Trabalho\Rateio_Regionais\exemplos")
    planilhas_encontradas = list(PASTA_BASE.glob("RATEIO PAG.xlsx"))
    if not planilhas_encontradas:
        print("⚠️ Planilha RATEIO PAG não encontrada na pasta!")
        return
    caminho_planilha_rateio = str(planilhas_encontradas[0].resolve())

    pdfs_encontrados = [arq for arq in PASTA_BASE.glob("*") if arq.suffix.lower() == ".pdf"]
    if not pdfs_encontrados:
        print("⚠️ Nenhum boleto PDF encontrado na pasta!")
        return
    elif len(pdfs_encontrados) > 1:
        print(f"⚠️ Atenção! Foram encontrados {len(pdfs_encontrados)} PDFs. Deixe apenas UM boleto na pasta!")
        return
    caminho_boleto_pdf = str(pdfs_encontrados[0].resolve())

    print(f"✅ Planilha de Upload carregada: {caminho_planilha_rateio}")
    print(f"✅ Boleto carregado: {caminho_boleto_pdf}")

    # 2. EXTRAÇÃO DE DADOS
    print("⏳ Extraindo dados do Boleto...")
    campos = extrair_campos_boleto(caminho_boleto_pdf)

    cnpj_correios = campos["cnpj_beneficiario"]
    num_doc       = campos["numero_documento"]
    vencimento    = campos["vencimento"]
    valor_boleto  = campos["valor_total_str"]
    cnpj_mrv      = campos["cnpj_pagador"]
    emissao_proc  = campos["data_processamento"]

    if not cnpj_mrv or not valor_boleto:
        print("⚠️ Não foi possível localizar o CNPJ da MRV ou o Valor no boleto.")
        return

    print(f"📌 Dados extraídos: CNPJ MRV: {cnpj_mrv} | Valor: R$ {valor_boleto} | Nº Doc: {num_doc}")

    ARQUIVO_REGRAS_XLSX = Path(__file__).with_name("dados_puxados_preenchimento.xlsx")
    df = pd.read_excel(ARQUIVO_REGRAS_XLSX, engine="openpyxl")

    texto_completo_pdf = norm_text(read_pdf_text(caminho_boleto_pdf)).upper()
    norm_limpo = texto_completo_pdf.replace(".", "").replace("/", "").replace("-", "")

    ID_REGIONAL = None
    candidatos = []

    for index, linha in df.iterrows():
        palavra_chave = str(linha.get("PALAVRA_CHAVE", "")).upper()
        if not palavra_chave or palavra_chave == "NAN": continue
        palavra_chave_limpa = palavra_chave.replace(".", "").replace("/", "").replace("-", "")
        if palavra_chave_limpa in norm_limpo:
            candidatos.append(linha)

    if len(candidatos) == 0:
        print("⚠️ Falha: Nenhuma 'PALAVRA_CHAVE' da planilha Excel foi encontrada no boleto.")
        return
    elif len(candidatos) == 1:
        linha_escolhida = candidatos[0]
        ID_REGIONAL  = linha_escolhida["ID"]
        descr        = linha_escolhida["DESCR"]
        material_cod = str(linha_escolhida["material_cod"])
    else:
        resultado = determinar_id_por_valor(valor_boleto, cnpj_mrv, df)
        if resultado is None:
            print("⚠️ Não foi possível determinar o ID regional pelo valor do boleto.")
            return
        ID_REGIONAL  = resultado["ID"]
        descr        = resultado["DESCR"]
        material_cod = resultado["material_cod"]

    print(f"📋 Regional: {ID_REGIONAL} | Descrição: {descr} | Material: {material_cod}")

    # 3. INICIALIZAÇÃO DO SELENIUM
    try:
        chrome_options = webdriver.ChromeOptions()
        driver = webdriver.Chrome(options=chrome_options) 
        driver.get("https://mrvpag2.mrv.com.br/home")
        driver.maximize_window()
        wait = WebDriverWait(driver, 15)
        wait_rapido = WebDriverWait(driver, 2)
    except Exception as e:
        print(f"Erro ao iniciar o Chrome: {e}")
        return

    # 4. NAVEGAÇÃO DO ROBÔ
    try:    
        print("Aguardando login...")
        wait.until(EC.presence_of_element_located((By.ID, "i0116"))).send_keys("pedro.henrsilva@mrv.com.br") # Digite seu e-mail
        click_anti_stale(wait, By.ID, "idSIButton9")
        wait.until(EC.presence_of_element_located((By.ID, "i0118"))).send_keys("Felipe22/") # Digite sua senha
        click_anti_stale(wait, By.ID, "idSIButton9")
        print("!!! APROVE O MFA NO CELULAR !!!")
        click_anti_stale(wait, By.ID, "idSIButton9") 
        
        fechar_mensagem = WebDriverWait(driver, 100).until(EC.element_to_be_clickable((By.CSS_SELECTOR, "mat-icon.btnCancelTest")))
        fechar_mensagem.click()
        print("Mensagem de boas-vindas fechada.")

        novo_protocolo = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "a.item-menu[routerlink='/protocolo'], a.item-menu[href='/protocolo']")))
        novo_protocolo.click()
        
        tipo_nota = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "mat-card a.pointer")))
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", tipo_nota)
        tipo_nota.click()
        
        tipo_de_documento = wait.until(EC.presence_of_element_located((By.XPATH, "//mat-expansion-panel-header[.//mat-panel-title[contains(normalize-space(),'Tipo de Documento')]]")))
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", tipo_de_documento)
        if tipo_de_documento.get_attribute("aria-expanded") == "false":
            wait.until(EC.element_to_be_clickable(tipo_de_documento)).click()
        
        select_el = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "mat-select[formcontrolname='frmTipoDoc'], mat-select[aria-label='Qual o tipo do documento'], #mat-select-2")))
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", select_el)
        select_el.click()

        wait.until(EC.visibility_of_element_located((By.CSS_SELECTOR, ".cdk-overlay-pane .mat-select-panel")))
        opcao = wait.until(EC.element_to_be_clickable((By.XPATH, "//mat-option[.//span[contains(@class,'mat-option-text') and normalize-space()='NF somente de Serviços']]")))
        driver.execute_script("arguments[0].scrollIntoView({block:'center'});", opcao)
        opcao.click()

        input_enviar_arquivo = WebDriverWait(driver, 20).until(EC.presence_of_element_located((By.CSS_SELECTOR, "#frmFile")))
        input_enviar_arquivo.send_keys(caminho_boleto_pdf)

        btn_continuar = wait.until(EC.element_to_be_clickable((By.ID, "btnTipoDoc")))
        safe_click(driver, btn_continuar) # Corrigido para passar o driver

        aprovador_select = wait.until(EC.element_to_be_clickable((By.XPATH, "//mat-select[@aria-label='APROVADOR' or @placeholder='APROVADOR']")))
        safe_click(driver, aprovador_select) # Corrigido para passar o driver

        opcao_vanessa = wait.until(EC.element_to_be_clickable((By.XPATH, "//mat-option//span[normalize-space(.)='VANESSA DE BRITO RODRIGUES (VANESSA.BRODRIGUES)']/ancestor::mat-option")))
        safe_click(driver, opcao_vanessa) # Corrigido para passar o driver

        cnpj_input = driver.find_element(By.CSS_SELECTOR, "input[placeholder='CNPJ DA EMPRESA MRV'], input[aria-label='CNPJ DA EMPRESA MRV']")
        cnpj_input.clear()
        cnpj_input.send_keys(cnpj_mrv)
        cnpj_input.send_keys(Keys.TAB)

        xpath_linha_mrv = "(//tr[contains(@class,'mat-row')][.//td[contains(normalize-space(.),'MRV')]])[1]"
        linha_mrv = wait.until(EC.presence_of_element_located((By.XPATH, xpath_linha_mrv)))
        driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", linha_mrv)

        try:
            wait.until(EC.element_to_be_clickable((By.XPATH, xpath_linha_mrv)))
        except TimeoutException:
            pass 

        try:
            try:
                ActionChains(driver).move_to_element(linha_mrv).pause(0.1).click(linha_mrv).perform()
            except (ElementClickInterceptedException, StaleElementReferenceException):
                primeiro_td = linha_mrv.find_element(By.XPATH, ".//td[1]")
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", primeiro_td)
                try:
                    ActionChains(driver).move_to_element(primeiro_td).pause(0.1).click(primeiro_td).perform()
                except Exception:
                    driver.execute_script("arguments[0].click();", primeiro_td)
        except Exception as e:
            print("Erro ao clicar na linha MRV.")
            raise

        wait_overlay_gone(driver, wait)
        wait_no_overlay(driver, wait)
        
        for _ in range(3):
            try:
                inp_num_doc = get_input_by_formcontrol(driver, wait, "frmNumDocumento")
                type_safely(driver, wait, inp_num_doc, num_doc)
                if (inp_num_doc.get_attribute("value") or "").strip() == num_doc: break
            except StaleElementReferenceException: continue

        wait_no_overlay(driver, wait)
        for _ in range(3):
            try:
                inp_cnpj_cor = get_input_by_formcontrol(driver, wait, "frmCnpjFornecedor")
                editable = ensure_enabled_and_editable(driver, inp_cnpj_cor, allow_force=True)
                if editable: type_safely(driver, wait, inp_cnpj_cor, cnpj_correios)
                else: js_set_value_and_dispatch(driver, inp_cnpj_cor, cnpj_correios)
                if (re.sub(r"\D", "", inp_cnpj_cor.get_attribute("value") or "")) == cnpj_correios: break
            except StaleElementReferenceException: continue

        wait_no_overlay(driver, wait)
        for _ in range(3):
            try:
                inp_emissao = get_input_by_formcontrol(driver, wait, "frmDtEmissao")
                try: click_with_fallback(driver, inp_emissao)
                except Exception: pass
                inp_emissao.send_keys(Keys.ESCAPE)
                type_safely(driver, wait, inp_emissao, emissao_proc)
                if (inp_emissao.get_attribute("value") or "").strip() == emissao_proc: break
            except StaleElementReferenceException: continue

        wait_no_overlay(driver, wait)
        for _ in range(3):
            try:
                inp_venc = get_input_by_formcontrol(driver, wait, "frmVencimento")
                ensure_enabled_and_editable(driver, inp_venc, allow_force=True)
                try: click_with_fallback(driver, inp_venc)
                except Exception: pass
                inp_venc.send_keys(Keys.ESCAPE)
                type_safely(driver, wait, inp_venc, vencimento)
                if (inp_venc.get_attribute("value") or "").strip() == vencimento: break
            except StaleElementReferenceException: continue

        wait_no_overlay(driver, wait)
        for _ in range(3):
            try:
                inp_valor = get_input_by_formcontrol(driver, wait, "frmValorTotalNf")
                click_with_fallback(driver, inp_valor)
                inp_valor.send_keys(Keys.CONTROL, 'a', Keys.DELETE)
                for ch in valor_boleto:
                    inp_valor.send_keys(ch)
                    time.sleep(0.01)
                if (inp_valor.get_attribute("value") or "").strip(): break
            except StaleElementReferenceException: continue

        qtd_cliques = click_ok_confirm(driver, wait_rapido, timeout=3, max_tentativas=3)
        print(f"✅ {qtd_cliques} diálogo(s) confirmado(s) com sucesso.")

        campo_desc = driver.find_element(By.CSS_SELECTOR, 'input[formcontrolname="frmDescNota"]')
        campo_desc.clear()
        campo_desc.send_keys(descr)

        btn1 = click_primeiro_continuar(driver, wait_rapido)
        esperar_transicao_apos_primeiro(wait_rapido, btn1)
        click_segundo_continuar(driver, wait_rapido)

        btn_adicionar = wait_rapido.until(EC.element_to_be_clickable((By.XPATH, "//span[normalize-space(.)='Adicionar']/ancestor::button[1]")))
        btn_adicionar.click()

        preencher_codigo_material_ultima_linha(driver, wait_rapido, material_cod, timeout=10)
        click_pesquisar(driver, wait_rapido)

        selecionar_primeira_linha_checkbox(driver, wait, timeout=40, textos_para_verificar=["S2601", "5001382", "CNAE/NCM"], exigir_todos=False, clicar_mesmo_se_faltar=True)
        click_incluir_produtos(driver, wait_rapido)
        preencher_quantidade_e_valor(driver, wait_rapido, quantidade="1", valor_boleto=valor_boleto)
        abrir_select_justificativa(driver, wait_rapido)
        selecionar_opcao_justificativa_com_hover(driver, wait_rapido, texto_alvo="2 - Orientações do gestor/coordendor da área")
        click_continuar_proximo_ao_select(driver, wait_rapido)
        
        print("Anexando planilha em segundo plano (sem abrir janela do Windows)...")
        try:
            # 1. Pega TODOS os inputs de arquivo escondidos na página
            inputs_file = driver.find_elements(By.XPATH, "//input[@type='file']")
            
            if not inputs_file:
                print("⚠️ Nenhum input de arquivo encontrado na página!")
            else:
                # 2. O pulo do gato: pega o ÚLTIMO input da lista (o primeiro era o do PDF)
                input_planilha = inputs_file[-1]
                
                # 3. Força o elemento a ficar visível para o Selenium conseguir interagir
                driver.execute_script(
                    "arguments[0].style.display = 'block'; "
                    "arguments[0].style.visibility = 'visible'; "
                    "arguments[0].style.opacity = 1;", 
                    input_planilha
                )
                
                # 4. Envia o caminho do arquivo Excel
                input_planilha.send_keys(caminho_planilha_rateio)
                print("✅ Planilha anexada com sucesso via HTML!")
                
                time.sleep(2) # Pequena pausa para o site processar o arquivo
                
        except Exception as e:
            print(f"⚠️ Erro ao tentar anexar silenciosamente: {e}")
            
        total_ok = click_ok_confirm_repeatedly(driver, wait, max_clicks=3)
        print(f"[INFO] Botão OK clicado {total_ok} vez(es).")

        print("✅ Fluxo concluído com sucesso! O navegador permanecerá aberto para conferência.")

    except Exception as e:
        print(f"❌ Erro Crítico durante a execução: {e}")
        traceback.print_exc()
        try:
            driver.save_screenshot("erro_final.png")
            debug_dump(driver, "erro_final")
        except Exception:
            pass
 

# ==============================================================================
# 3. SUAS FUNÇÕES ORIGINAIS (Cole suas funções aqui)
# ==============================================================================


def click_anti_stale(wait, by, seletor, tentativas=3):
    """Procura o elemento e clica. Se ele virar 'fantasma' (Stale), procura de novo."""
    for _ in range(tentativas):
        try:
            # O segredo é SEMPRE procurar o elemento de novo dentro do 'try'
            elemento = wait.until(EC.element_to_be_clickable((by, seletor)))
            elemento.click()
            return True # Deu certo, sai da função
        except StaleElementReferenceException:
            time.sleep(0.5) # Espera a tela terminar de piscar e tenta de novo
    raise RuntimeError(f"O elemento {seletor} sumiu repetidas vezes.")

def scroll_center(driver, el):
    try: driver.execute_script("arguments[0].scrollIntoView({block:'center'});", el)
    except: pass

def wait_overlays_to_hide(wait):
    try: wait.until_not(lambda d: len(d.find_elements(By.CSS_SELECTOR, ".cdk-overlay-backdrop, .mat-progress-spinner")) > 0)
    except: pass

def safe_click(driver, element):
    # Garante que o elemento esteja visível e tenta click normal; se falhar, usa JS.
    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
    try:
        element.click()
    except Exception:
        driver.execute_script("arguments[0].click();", element)

def safe_click_diferenciado(driver, element):
    # Garante que o elemento esteja visível e tenta click normal; se falhar, usa JS.
    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
    try:
        element.click()
    except Exception:
        driver.execute_script("arguments[0].click();", element)
 
def fazer_upload_janela_windows(caminho_do_arquivo):
    """
    Interage com a janela nativa 'Abrir' do Windows usando pywinauto.
    Tenta múltiplos backends e títulos para máxima compatibilidade.
    """
    from pywinauto import Desktop
    from pywinauto.timings import wait_until
    import time

    caminho_absoluto = str(Path(caminho_do_arquivo).resolve())

    # Títulos possíveis da janela (pt-BR, en-US, variações do Chrome)
    titulos_possiveis = [
        "Abrir",
        "Open",
        "Abrir arquivo",
        "Escolher arquivo para carregar",
        "Escolher arquivo",
        "Select file",
        "File Upload",
        "Carregar",
    ]

    # Classes de janela conhecidas para diálogos de arquivo do Windows
    classes_possiveis = ["#32770"]  # Classe padrão de diálogos do Windows

    janela = None
    backend_usado = None

    # ── Tentativa 1: Busca por TÍTULO em ambos os backends ──
    for backend in ["uia", "win32"]:
        if janela:
            break
        try:
            desktop = Desktop(backend=backend)
            for titulo in titulos_possiveis:
                try:
                    win = desktop.window(title=titulo)
                    if win.exists(timeout=1):
                        janela = win
                        backend_usado = backend
                        print(f"✅ Janela encontrada! Título='{titulo}', Backend='{backend}'")
                        break
                except Exception:
                    continue
        except Exception:
            continue

    # ── Tentativa 2: Busca por TÍTULO PARCIAL (title_re) ──
    if not janela:
        import re
        for backend in ["uia", "win32"]:
            if janela:
                break
            try:
                desktop = Desktop(backend=backend)
                # Busca qualquer janela que contenha "Abrir" ou "Open" no título
                for win in desktop.windows():
                    try:
                        titulo = win.window_text()
                        if any(t.lower() in titulo.lower() for t in ["abrir", "open", "upload", "file"]):
                            janela = win
                            backend_usado = backend
                            print(f"✅ Janela encontrada (parcial)! Título='{titulo}', Backend='{backend}'")
                            break
                    except Exception:
                        continue
            except Exception:
                continue

    # ── Tentativa 3: Busca pela CLASSE da janela (#32770 = diálogo padrão Windows) ──
    if not janela:
        for backend in ["uia", "win32"]:
            if janela:
                break
            try:
                desktop = Desktop(backend=backend)
                for classe in classes_possiveis:
                    try:
                        win = desktop.window(class_name=classe)
                        if win.exists(timeout=1):
                            janela = win
                            backend_usado = backend
                            titulo_real = win.window_text()
                            print(f"✅ Janela encontrada (por classe)! Título='{titulo_real}', Classe='{classe}', Backend='{backend}'")
                            break
                    except Exception:
                        continue
            except Exception:
                continue

    # ── Se nenhuma tentativa funcionou ──
    if not janela:
        # DEBUG: Lista todas as janelas para diagnóstico
        print("\n⚠️ JANELA NÃO ENCONTRADA! Listando todas as janelas abertas:")
        for backend in ["uia", "win32"]:
            try:
                desktop = Desktop(backend=backend)
                print(f"\n  --- Backend {backend} ---")
                for win in desktop.windows():
                    try:
                        t = win.window_text()
                        c = win.class_name()
                        if t:
                            print(f"    Título: '{t}' | Classe: '{c}'")
                    except Exception:
                        pass
            except Exception:
                pass
        raise RuntimeError(
            "⚠️ Janela de upload não encontrada! "
            "Verifique se a janela 'Abrir' está realmente visível na tela."
        )

    # ══════════════════════════════════════════════════════════════
    # INTERAÇÃO COM A JANELA ENCONTRADA
    # ══════════════════════════════════════════════════════════════

    try:
        janela.wait("ready", timeout=10)
    except Exception:
        time.sleep(1)  # Fallback: espera simples

    # ── Preencher o campo "Nome do arquivo" ──
    campo_preenchido = False

    if backend_usado == "uia":
        # UIA: tenta localizar o Edit pelo automation_id ou pelo título
        tentativas_campo = [
            lambda: janela.child_window(title="Nome do arquivo:", control_type="ComboBox").child_window(control_type="Edit"),
            lambda: janela.child_window(title="Nome do arquivo:", control_type="Edit"),
            lambda: janela.child_window(title="File name:", control_type="ComboBox").child_window(control_type="Edit"),
            lambda: janela.child_window(title="File name:", control_type="Edit"),
            lambda: janela.child_window(control_type="Edit", found_index=0),
        ]
    else:
        # win32: tenta pelo class_name do Edit
        tentativas_campo = [
            lambda: janela.child_window(class_name="Edit", found_index=0),
            lambda: janela.child_window(title="Nome do arquivo:", class_name="ComboBoxEx32").child_window(class_name="Edit"),
            lambda: janela.child_window(class_name="ComboBoxEx32").child_window(class_name="Edit"),
        ]

    for get_campo in tentativas_campo:
        try:
            campo = get_campo()
            if campo.exists(timeout=2):
                campo.set_edit_text(caminho_absoluto)
                campo_preenchido = True
                print(f"✅ Caminho digitado no campo: {caminho_absoluto}")
                break
        except Exception:
            continue

    if not campo_preenchido:
        # Fallback FINAL: usa keyboard do pywinauto para digitar
        try:
            from pywinauto.keyboard import send_keys
            janela.set_focus()
            time.sleep(0.3)
            # Limpa o campo atual e digita o caminho
            send_keys("^a{DELETE}", pause=0.05)
            time.sleep(0.1)
            # Escapa caracteres especiais do pywinauto ({, }, +, ^, %)
            caminho_escaped = caminho_absoluto.replace("{", "{{").replace("}", "}}")
            send_keys(caminho_escaped, pause=0.02, with_spaces=True)
            campo_preenchido = True
            print(f"✅ Caminho digitado via send_keys: {caminho_absoluto}")
        except Exception as e:
            raise RuntimeError(f"⚠️ Não consegui digitar o caminho do arquivo: {e}")

    time.sleep(0.5)

    # ── Clicar no botão "Abrir" ──
    botao_clicado = False

    nomes_botao = ["Abrir", "Open", "&Abrir", "&Open"]

    for nome_btn in nomes_botao:
        if botao_clicado:
            break
        try:
            if backend_usado == "uia":
                btn = janela.child_window(title=nome_btn, control_type="Button")
            else:
                btn = janela.child_window(title=nome_btn, class_name="Button")

            if btn.exists(timeout=2):
                btn.click()
                botao_clicado = True
                print(f"✅ Botão '{nome_btn}' clicado!")
        except Exception:
            continue

    if not botao_clicado:
        # Fallback: pressiona Enter
        try:
            from pywinauto.keyboard import send_keys
            send_keys("{ENTER}")
            botao_clicado = True
            print("✅ Enter pressionado como fallback para confirmar.")
        except Exception as e:
            raise RuntimeError(f"⚠️ Não consegui clicar no botão Abrir: {e}")

    # ── Aguarda a janela fechar ──
    for _ in range(20):
        try:
            if not janela.exists(timeout=0):
                print(f"✅ Upload concluído: {caminho_absoluto}")
                return
        except Exception:
            print(f"✅ Upload concluído: {caminho_absoluto}")
            return
        time.sleep(0.5)

    print("⚠️ Janela pode não ter fechado, mas o caminho foi preenchido.")

def click_ok_confirm_repeatedly(driver, wait, max_clicks=5):    
    clicks = 0
    locators = [
        (By.ID, "btnTipoDoc"),
        (By.XPATH, "//button[contains(@class,'confirm') and normalize-space(.)='OK']")
    ]
    for _ in range(max_clicks):
        wait_overlays_to_hide(wait)
        btn = None
        for by, sel in locators:
            try:
                elem = driver.find_element(by, sel)
                if elem.is_displayed(): btn = elem; break
            except: pass
        if not btn: break
        
        safe_click(driver, btn)
        clicks += 1
        time.sleep(1)
    return clicks

def click_anexar_planilha(driver, wait, caminho_planilha=None):
    wait_overlays_to_hide(wait)

    locator_btn = (
        By.XPATH,
        "//button[.//span[normalize-space(.)='ANEXAR PLANILHA'] "
        " and not(@disabled) and not(contains(@class,'mat-button-disabled'))]"
    )

    try:
        btn = wait.until(EC.presence_of_element_located(locator_btn))
    except TimeoutException:
        debug_dump(driver, "anexar_planilha_nao_presente")
        raise RuntimeError("Botão 'ANEXAR PLANILHA' não encontrado.")

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
    time.sleep(0.05)

    try:
        wait.until(EC.element_to_be_clickable(locator_btn))
    except TimeoutException:
        print("[INFO] 'ANEXAR PLANILHA' não confirmou clicável; tentando assim mesmo.")

    # Clique no botão
    try:
        if not is_center_clickable_js(driver, btn):
            driver.execute_script("window.scrollBy(0, -80);")
            time.sleep(0.05)
        btn.click()
    except Exception as e1:
        print(f"[INFO] Clique direto (ANEXAR PLANILHA) falhou: {repr(e1)}")
        try:
            ActionChains(driver).move_to_element(btn).pause(0.05).click().perform()
        except Exception as e2:
            print(f"[INFO] ActionChains (ANEXAR PLANILHA) falhou: {repr(e2)}")
            try:
                driver.execute_script("arguments[0].click();", btn)
            except Exception as e3:
                debug_dump(driver, "anexar_planilha_click_fail")
                raise RuntimeError(f"Falha ao clicar em 'ANEXAR PLANILHA': {repr(e3)}")

    # (Opcional) Se existir input[type=file], envia o arquivo
    if caminho_planilha:
        # Aguarda input de arquivo ficar presente (às vezes é criado após o clique)
        try:
            input_file = wait.until(EC.presence_of_element_located((By.XPATH, "//input[@type='file']")))
        except TimeoutException:
            print("[INFO] Nenhum input[type=file] detectado após o clique; talvez abra um diálogo nativo.")
            return

        # Garante que o path é absoluto
        caminho_planilha = str(Path(caminho_planilha).resolve())

        try:
            input_file.send_keys(caminho_planilha)
        except Exception as e:
            # Se estiver oculto, tenta forçar visibilidade e reenviar
            try:
                driver.execute_script("arguments[0].style.display='block'; arguments[0].style.visibility='visible';", input_file)
                time.sleep(0.05)
                input_file.send_keys(caminho_planilha)
            except Exception as e2:
                debug_dump(driver, "anexar_planilha_upload_fail")
                raise RuntimeError(f"Falha ao enviar arquivo para input[type=file]: {repr(e2)}")

def wait_overlay_gone(driver, wait, timeout=40):
    try:
        WebDriverWait(driver, timeout).until(
            EC.invisibility_of_element_located((
                By.CSS_SELECTOR,
                ".cdk-overlay-backdrop.cdk-overlay-backdrop-showing, "
                ".mat-progress-bar, .mat-spinner, .ngx-spinner-overlay"
            ))
        )
    except TimeoutException:
        pass

def get_visible_input(driver, formcontrol: str):
    candidates = driver.find_elements(
        By.CSS_SELECTOR, f"input[formcontrolname='{formcontrol}']"
    )
    visibles = [el for el in candidates if el.is_displayed() and el.is_enabled()]
    if not visibles:
        raise TimeoutException(f"Input visível '{formcontrol}' não encontrado.")
    return visibles[0]

def focus_input(driver, wait, el):
    try:
        driver.execute_script(
            "arguments[0].scrollIntoView({block:'center'});", el
        )
        wait.until(lambda d: el.is_displayed() and el.is_enabled())
        try:
            el.click()
        except ElementClickInterceptedException:
            container = el.find_element(
                By.XPATH,
                "./ancestor::mat-form-field//div[contains(@class,'mat-form-field-flex')]"
            )
            driver.execute_script(
                "arguments[0].scrollIntoView({block:'center'});", container
            )
            try:
                ActionChains(driver).move_to_element(container).pause(0.05).click(container).perform()
            except Exception:
                driver.execute_script("arguments[0].click();", container)
    except StaleElementReferenceException:
        pass

def clear_and_type(el, text: str):
    """Versão unificada — tenta Selenium, fallback JS."""
    driver = el.parent  # Funciona com qualquer WebElement

    try:
        driver.execute_script(
            "arguments[0].scrollIntoView({block:'center'});", el
        )
        time.sleep(0.15)
    except Exception:
        pass

    # Foca o elemento
    try:
        el.click()
    except Exception:
        try:
            driver.execute_script("arguments[0].click(); arguments[0].focus();", el)
        except Exception:
            pass

    # Limpa e digita
    try:
        el.send_keys(Keys.CONTROL, 'a', Keys.DELETE)
        time.sleep(0.1)
        if text:
            el.send_keys(str(text))
    except Exception:
        # Fallback JS
        driver.execute_script("""
            const el = arguments[0], val = arguments[1] ?? '';
            el.value = '';
            el.dispatchEvent(new Event('input', {bubbles: true}));
            el.value = val;
            el.dispatchEvent(new Event('input', {bubbles: true}));
            el.dispatchEvent(new Event('change', {bubbles: true}));
        """, el, text or '')

def fill_input(driver, wait, formcontrol: str, value: str, numeric=False):
    wait_overlay_gone(driver, wait)
    el = get_visible_input(driver, formcontrol)
    focus_input(driver, wait, el)
    if numeric and value is not None:
        value = re.sub(r'\D', '', str(value))
    clear_and_type(el, value)
    try:
        WebDriverWait(driver, 5).until(
            lambda d: (el.get_attribute("value") or "") != ""
        )
    except Exception:
        pass


def somente_digitos(s: str) -> str:
    return re.sub(r"\D", "", s or "")

def norm_text(s: str) -> str:
    s = unicodedata.normalize("NFKD", s).encode("ASCII", "ignore").decode("ASCII")
    s = s.replace("\xa0", " ")
    return re.sub(r"[ \t]+", " ", s)

def read_pdf_text(pdf_path: str) -> str:
    reader = PdfReader(pdf_path)
    txt = [p.extract_text() or "" for p in reader.pages]
    texto = "\n".join(txt)
    if not texto.strip(): raise ValueError("PDF sem texto. Use OCR.")
    return texto

def extrair_cnpj_pagador(text_norm: str) -> Optional[str]:
    # Extrai o CNPJ da MRV ignorando os CNPJs dos Correios
    padrao_cnpj = r"\b\d{2}\.\d{3}\.\d{3}/\d{4}-\d{2}\b"
    cnpjs_encontrados = re.findall(padrao_cnpj, text_norm)
    cnpjs_correios = ["34.028.316/0015-09", "34.028.316/0001-03"]
    for cnpj in cnpjs_encontrados:
        if cnpj not in cnpjs_correios:
            return cnpj
    return None

def extrair_numero_documento_7d(text_norm: str) -> Optional[str]:
    mdoc = re.search(r"DOCUMENTO.{0,100}(\d{7})", text_norm, flags=re.I | re.S)
    return mdoc.group(1) if mdoc else None

def extrair_valor_total(text_norm: str) -> Optional[str]:
    anchor = r"(?i)VALOR\s*(?:DO\s*)?DOCUMENTO(?:\s*\(R\$\))?"
    money_re = r"(?:R\$\s*)?(\d{1,3}(?:\.\d{3})*,\d{2})"
    lines = text_norm.splitlines()
    for i, line in enumerate(lines):
        if re.search(anchor, line, flags=re.I):
            for j in (i, i+1, i+2): # Busca na mesma linha ou nas duas de baixo
                if 0 <= j < len(lines):
                    m = re.search(money_re, lines[j], flags=re.I)
                    if m: return m.group(1)
    return None

def extrair_datas_correios(text_norm: str) -> dict:
    """
    Extrai data de emissão e vencimento do boleto dos Correios.

    Regra de negócio (infalível):
      - Vencimento = MAIOR data encontrada no documento
      - Emissão    = SEGUNDA data única (ordenada), ou seja,
                     a segunda após a menor data encontrada

    Mesmo que os rótulos do regex acertem, a validação final
    garante que vencimento >= emissão.
    """
    from datetime import datetime

    # ── DEBUG ──
    print("=" * 60)
    print("🔍 DEBUG — Texto normalizado do PDF (primeiras 2000 chars):")
    print(text_norm[:2000])
    print("=" * 60)

    # ── Passo 1: Extrair TODAS as datas do documento ──
    todas_datas_str = re.findall(r"\d{2}/\d{2}/\d{4}", text_norm)
    print(f"  🔎 Todas as datas encontradas (com repetições): {todas_datas_str}")

    # Remove duplicatas mantendo a ordem de aparição
    vistas = set()
    unicas_str = []
    for d in todas_datas_str:
        if d not in vistas:
            vistas.add(d)
            unicas_str.append(d)

    # Converte para datetime e ordena cronologicamente
    datas_parsed = []
    for d_str in unicas_str:
        try:
            dt = datetime.strptime(d_str, "%d/%m/%Y")
            datas_parsed.append((d_str, dt))
        except ValueError:
            continue  # Ignora datas inválidas (ex: 32/13/2025)

    datas_parsed.sort(key=lambda x: x[1])

    print(f"  📋 Datas únicas ordenadas: {[d[0] for d in datas_parsed]}")

    if not datas_parsed:
        print("  ⚠️ Nenhuma data válida encontrada no PDF!")
        return {"emissao": "", "vencimento": ""}

    # ── Passo 2: Tentar extrair por rótulos (para log/debug) ──
    emissao_rotulo = ""
    vencimento_rotulo = ""

    rotulos_emissao = [
        r"Data\s*(?:de\s*)?Processamento[^\d]{0,30}(\d{2}/\d{2}/\d{4})",
        r"Data\s*(?:do\s*)?Documento[^\d]{0,30}(\d{2}/\d{2}/\d{4})",
        r"Dt\.?\s*Processamento[^\d]{0,30}(\d{2}/\d{2}/\d{4})",
    ]
    rotulos_vencimento = [
        r"Vencimento[^\d]{0,30}(\d{2}/\d{2}/\d{4})",
        r"Dt\.?\s*Vencimento[^\d]{0,30}(\d{2}/\d{2}/\d{4})",
        r"Data\s*Vencimento[^\d]{0,30}(\d{2}/\d{2}/\d{4})",
    ]

    for padrao in rotulos_emissao:
        m = re.search(padrao, text_norm, re.IGNORECASE)
        if m:
            emissao_rotulo = m.group(1)
            print(f"  📅 Rótulo emissão capturou: {emissao_rotulo}")
            break

    for padrao in rotulos_vencimento:
        m = re.search(padrao, text_norm, re.IGNORECASE)
        if m:
            vencimento_rotulo = m.group(1)
            print(f"  📅 Rótulo vencimento capturou: {vencimento_rotulo}")
            break

    # ── Passo 3: Aplicar a REGRA DE NEGÓCIO (sempre prevalece) ──
    #
    #   Datas ordenadas: [menor, ..., ..., maior]
    #
    #   Vencimento = MAIOR data (última na lista ordenada)
    #   Emissão    = SEGUNDA data única (índice 1 da lista ordenada)
    #               Se só houver 2 datas, emissão = a menor (índice 0)
    #               Se só houver 1 data, emissão = vencimento = essa data

    vencimento = datas_parsed[-1][0]  # Sempre a MAIOR

    if len(datas_parsed) >= 3:
        # 3+ datas: menor é algo irrelevante (ex: data do cedente),
        # segunda é a emissão, última é o vencimento
        emissao = datas_parsed[1][0]
    elif len(datas_parsed) == 2:
        # 2 datas: menor = emissão, maior = vencimento
        emissao = datas_parsed[0][0]
    else:
        # 1 data: usa a mesma para ambos
        emissao = datas_parsed[0][0]

    # ── Passo 4: Validação cruzada com rótulos (log de divergência) ──
    if emissao_rotulo and emissao_rotulo != emissao:
        print(f"  ⚠️ DIVERGÊNCIA: Rótulo emissão={emissao_rotulo}, Regra negócio={emissao}")
        print(f"     → Usando regra de negócio: {emissao}")

    if vencimento_rotulo and vencimento_rotulo != vencimento:
        print(f"  ⚠️ DIVERGÊNCIA: Rótulo vencimento={vencimento_rotulo}, Regra negócio={vencimento}")
        print(f"     → Usando regra de negócio: {vencimento}")

    # ── Passo 5: Validação final de sanidade ──
    dt_emissao = datetime.strptime(emissao, "%d/%m/%Y")
    dt_vencimento = datetime.strptime(vencimento, "%d/%m/%Y")

    if dt_vencimento < dt_emissao:
        # Isso NUNCA deveria acontecer com a lógica acima, mas por segurança:
        print(f"  🚨 ERRO LÓGICO: vencimento ({vencimento}) < emissão ({emissao}). Invertendo!")
        emissao, vencimento = vencimento, emissao

    print(f"  ✅ RESULTADO FINAL → Emissão: {emissao} | Vencimento: {vencimento}")
    return {"emissao": emissao, "vencimento": vencimento}

def extrair_campos_boleto(pdf_path: str) -> Dict[str, Optional[str]]:
    texto = read_pdf_text(pdf_path)
    norm  = norm_text(texto)

    num_doc = extrair_numero_documento_7d(norm)
    datas = extrair_datas_correios(norm) # Chama a nova inteligência de datas
    valor_total_str = extrair_valor_total(norm)
    cnpj_pagador_extraido = extrair_cnpj_pagador(norm)
    
    cnpj_pag = somente_digitos(cnpj_pagador_extraido) if cnpj_pagador_extraido else None
    
    return {
        "numero_documento": num_doc,
        "cnpj_pagador": cnpj_pag,
        "cnpj_beneficiario": CNPJ_CORREIOS_FIXO,
        "data_processamento": datas["emissao"],
        "vencimento": datas["vencimento"], # 🔹 AGORA O VENCIMENTO VOLTOU!
        "valor_total_str": valor_total_str     
    }

def determinar_id_por_valor(valor_str: str, cnpj_pagador: str, df: pd.DataFrame) -> dict:
    """
    Quando o CNPJ do pagador aparece em múltiplas linhas da planilha,
    usa o valor do boleto para determinar qual ID/regional é o correto.

    Regras:
      - Valor < 400.00        → ID = 2  (Cartões Uberlândia)
      - 400.00 ≤ Valor < 2000 → ID = 7  (Centro Oeste Brasília)
      - Valor ≥ 40000         → ID = 8  (Belo Horizonte AGF)
    """
    # Converte o valor string para float (ex: "1.234,56" → 1234.56)
    valor_float = float(
        valor_str
        .replace(".", "")   # remove separador de milhar
        .replace(",", ".")  # troca vírgula decimal por ponto
    )

    # Filtra as linhas da planilha que correspondem ao CNPJ
    cnpj_limpo = cnpj_pagador.replace(".", "").replace("/", "").replace("-", "")
    df["PALAVRA_CHAVE_LIMPA"] = (
        df["PALAVRA_CHAVE"]
        .astype(str)
        .str.replace(".", "", regex=False)
        .str.replace("/", "", regex=False)
        .str.replace("-", "", regex=False)
        .str.upper()
    )

    linhas_cnpj = df[df["PALAVRA_CHAVE_LIMPA"] == cnpj_limpo]

    if linhas_cnpj.empty:
        return None

    # Se só tem uma linha para esse CNPJ, retorna direto
    if len(linhas_cnpj) == 1:
        linha = linhas_cnpj.iloc[0]
        return {
            "ID": linha["ID"],
            "DESCR": linha["DESCR"],
            "material_cod": str(linha["material_cod"]),
        }

    # Múltiplas linhas → aplica regra de valor
    if valor_float < 400.00:
        id_alvo = 2
    elif valor_float < 2000.00:
        id_alvo = 7
    elif valor_float >= 40000.00:
        id_alvo = 8
    else:
        # Faixa entre 2000 e 40000 — ajuste conforme sua regra de negócio
        # Por ora, levanta erro para você mapear
        raise ValueError(
            f"⚠️ Valor R$ {valor_str} ({valor_float}) não se encaixa nas faixas "
            f"definidas (< 400 → ID 2 | 400–2000 → ID 7 | ≥ 40000 → ID 8)."
        )

    # Busca a linha com o ID determinado
    linha_alvo = linhas_cnpj[linhas_cnpj["ID"] == id_alvo]

    if linha_alvo.empty:
        raise ValueError(
            f"⚠️ ID {id_alvo} (determinado pelo valor R$ {valor_str}) "
            f"não encontrado na planilha para o CNPJ {cnpj_pagador}."
        )

    linha = linha_alvo.iloc[0]
    return {
        "ID": linha["ID"],
        "DESCR": linha["DESCR"],
        "material_cod": str(linha["material_cod"]),
    }

def click_continuar_proximo_ao_select(driver, wait):
    # Container do mat-select justificativa
    container = driver.find_element(
        By.XPATH,
        "//mat-form-field[.//mat-select[@formcontrolname='justificativa']]"
    )
    # Pega o primeiro link CONTINUAR que aparece depois desse container
    locator_local = (
        By.XPATH,
        ".//following::a[.//span[normalize-space(.)='CONTINUAR'] "
        " and (not(@aria-disabled) or @aria-disabled='false') "
        " and not(contains(@class,'mat-button-disabled'))][1]"
    )
    try:
        link = container.find_element(*locator_local)
    except Exception:
        raise RuntimeError("Não achei o CONTINUAR referente a este passo (após o select).")

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", link)
    time.sleep(0.05)

    from selenium.webdriver.support import expected_conditions as EC
    try:
        wait.until(EC.element_to_be_clickable((By.XPATH, "//a[.//span[normalize-space(.)='CONTINUAR']]")))
    except TimeoutException:
        pass

    # Clique normal → ActionChains → JS
    try:
        link.click()
    except Exception:
        try:
            ActionChains(driver).move_to_element(link).pause(0.05).click().perform()
        except Exception:
            driver.execute_script("arguments[0].click();", link)

def preencher_codigo_material_ultima_linha(driver, wait, material_cod, timeout=3):
    # Seleciona APENAS inputs de Código do Material (nunca pega CNAE/NCM)
    xpath_all = ("//input[(@formcontrolname='frmCodigoMaterial' or @name='codigoMaterial' "
                 " or @placeholder='CÓDIGO DO MATERIAL') and not(@disabled)]")

    # Espera pelo menos 1 visível 
    end = time.time() + timeout
    visiveis = []
    while time.time() < end:
        elems = driver.find_elements(By.XPATH, xpath_all)
        visiveis = [e for e in elems if e.is_displayed()]
        if visiveis:
            break
        time.sleep(0.2)

    if not visiveis:
        debug_dump(driver, "codigo_material_nenhum_visivel")
        raise RuntimeError("Nenhum campo 'CÓDIGO DO MATERIAL' visível encontrado.")

    # Pega o ÚLTIMO visível (tipicamente a 2ª linha recém-renderizada)
    alvo = visiveis[-1]

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", alvo)
    time.sleep(0.05)
    try:
        wait.until(EC.element_to_be_clickable((By.XPATH, xpath_all)))
    except TimeoutException:
        pass

    try:
        alvo.click()
    except Exception:
        driver.execute_script("arguments[0].focus();", alvo)

    clear_and_type(alvo, material_cod)

def click_incluir_produtos(driver, wait):
    wait_overlays_to_hide(wait)

    locator_btn = (
        By.XPATH,
        "//mat-action-row//button[.//span[normalize-space(.)='INCLUIR PRODUTO(S)'] "
        " and not(@disabled) and not(contains(@class,'mat-button-disabled'))]"
    )

    try:
        btn = wait.until(EC.presence_of_element_located(locator_btn))
    except TimeoutException:
        debug_dump(driver, "incluir_produtos_nao_presente")
        raise RuntimeError("Botão 'INCLUIR PRODUTO(S)' não encontrado.")

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
    time.sleep(0.05)

    try:
        wait.until(EC.element_to_be_clickable(locator_btn))
    except TimeoutException:
        print("[INFO] 'INCLUIR PRODUTO(S)' não confirmou clicável; tentando assim mesmo.")

    try:
        if not is_center_clickable_js(driver, btn):
            driver.execute_script("window.scrollBy(0, -80);")
            time.sleep(0.05)
        btn.click()
        return
    except Exception as e1:
        print(f"[INFO] Clique direto (INCLUIR PRODUTO(S)) falhou: {repr(e1)}")

    try:
        ActionChains(driver).move_to_element(btn).pause(0.05).click().perform()
        return
    except Exception as e2:
        print(f"[INFO] ActionChains (INCLUIR PRODUTO(S)) falhou: {repr(e2)}")

    try:
        driver.execute_script("arguments[0].click();", btn)
        return
    except Exception as e3:
        debug_dump(driver, "incluir_produtos_click_fail")
        raise RuntimeError(f"Falha ao clicar em 'INCLUIR PRODUTO(S)': {repr(e3)}")

def abrir_select_justificativa(driver, wait):
    wait_overlays_to_hide(wait)

    # Localizadores do mat-select (preferência pelo formcontrolname)
    loc_select = (
        By.CSS_SELECTOR,
        "mat-select[formcontrolname='justificativa']"
    )

    try:
        sel = wait.until(EC.presence_of_element_located(loc_select))
    except TimeoutException:
        # Fallback por aria-label/placeholder
        try:
            sel = wait.until(EC.presence_of_element_located((
                By.XPATH,
                "//mat-select[@formcontrolname='justificativa' or @aria-label='Por quê o Pedido não foi criado antes da emissão da Nota Fiscal?' or @placeholder='Por quê o Pedido não foi criado antes da emissão da Nota Fiscal?']"
            )))
        except TimeoutException:
            debug_dump(driver, "select_justificativa_nao_encontrado")
            raise RuntimeError("Campo select 'Justificativa' não encontrado.")

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", sel)
    time.sleep(0.05)

    # Tenta abrir o painel
    try:
        sel.click()
    except Exception as e1:
        print(f"[INFO] Click direto no mat-select falhou: {repr(e1)}")
        try:
            # Click na trigger interna (às vezes é mais confiável)
            trigger = sel.find_element(By.CSS_SELECTOR, ".mat-select-trigger")
            trigger.click()
        except Exception as e2:
            print(f"[INFO] Click na trigger falhou: {repr(e2)}")
            try:
                driver.execute_script("arguments[0].click();", sel)
            except Exception as e3:
                debug_dump(driver, "select_justificativa_click_fail")
                raise RuntimeError(f"Falha ao abrir o select 'Justificativa': {repr(e3)}")

    # Aguarda painel de opções aparecer
    try:
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, ".cdk-overlay-pane .mat-select-panel")))
    except TimeoutException:
        debug_dump(driver, "select_justificativa_sem_painel")
        raise RuntimeError("Painel de opções do 'Justificativa' não apareceu.")

def verificar_textos_na_tabela(driver, wait, textos, timeout=3):
    """
    Verifica quais textos de uma lista estão presentes no texto renderizado da tabela.

    Retorna:
      - encontrados (list): textos encontrados
      - faltando (list): textos não encontrados
      - conteudo (str): texto completo normalizado da tabela (útil para debug)
    """
    locator_container = (By.CSS_SELECTOR, "div.table-container")

    try:
        container = wait.until(EC.presence_of_element_located(locator_container))
    except TimeoutException:
        debug_dump(driver, "table_container_nao_presente")
        raise RuntimeError("Não encontrei o container da tabela (div.table-container).")

    # Normaliza espaços/quebras para bater melhor com o que “aparece na tela”
    conteudo = " ".join(container.text.split())

    encontrados = [t for t in textos if t and t in conteudo]
    faltando = [t for t in textos if t and t not in conteudo]

    return encontrados, faltando, conteudo

def click_ok_confirm(driver, wait, timeout=3, max_tentativas=3):
    """
    Clica no botão 'OK' (id='btnTipoDoc', class='confirm') dentro de
    diálogos Angular Material (mat-dialog-container).

    Lida com:
      - O botão estar DENTRO de um overlay (não espera overlay sumir antes)
      - O mesmo diálogo aparecer múltiplas vezes consecutivas
      - Elementos stale (diálogo re-renderizado)

    Retorna o número de vezes que clicou em OK.
    """
    cliques = 0

    for tentativa in range(max_tentativas):
        # ── 1) Esperar que um mat-dialog-container apareça ──
        locator_dialog = (
            By.CSS_SELECTOR,
            "mat-dialog-container"
        )

        try:
            WebDriverWait(
                driver, timeout if tentativa == 0 else 5,
                poll_frequency=0.3,
                ignored_exceptions=[StaleElementReferenceException],
            ).until(EC.presence_of_element_located(locator_dialog))
        except TimeoutException:
            if cliques == 0:
                print(f"[WARN] Nenhum diálogo mat-dialog-container apareceu após {timeout}s.")
                return 0
            else:
                # Já clicou antes, não apareceu de novo → acabou
                print(f"[INFO] Nenhum novo diálogo apareceu. Total de cliques OK: {cliques}")
                return cliques

        # ── 2) Localizar o botão OK DENTRO do diálogo ──
        locators_btn = [
            # Dentro do mat-dialog-container específico
            (By.CSS_SELECTOR, "mat-dialog-container button#btnTipoDoc"),
            (By.CSS_SELECTOR, "mat-dialog-container button.confirm"),
            (By.XPATH, "//mat-dialog-container//button[normalize-space()='OK']"),
        ]

        btn = None
        for by, sel in locators_btn:
            try:
                # Pega todos e filtra os visíveis
                elementos = driver.find_elements(by, sel)
                visiveis = [e for e in elementos if _is_displayed_safe(e)]

                if visiveis:
                    btn = visiveis[-1]  # Último visível (mais recente)
                    break
            except Exception:
                continue

        if btn is None:
            # Botão não encontrado, mas o diálogo existe → tenta esperar um pouco
            try:
                btn = WebDriverWait(
                    driver, 5,
                    poll_frequency=0.2,
                    ignored_exceptions=[StaleElementReferenceException],
                ).until(EC.element_to_be_clickable((
                    By.CSS_SELECTOR,
                    "mat-dialog-container button#btnTipoDoc"
                )))
            except TimeoutException:
                print(f"[WARN] Diálogo presente mas botão OK não encontrado (tentativa {tentativa + 1}).")
                if cliques > 0:
                    return cliques
                continue

        # ── 3) Scroll + Clique com fallback ──
        try:
            driver.execute_script(
                "arguments[0].scrollIntoView({block:'center'});", btn
            )
            time.sleep(0.1)
        except StaleElementReferenceException:
            continue  # Diálogo sumiu/re-renderizou, tenta de novo

        metodo = _clicar_botao_ok(driver, btn)
        if metodo:
            cliques += 1
            print(f"[OK] Clique #{cliques} no botão OK via {metodo} (tentativa {tentativa + 1})")
        else:
            print(f"[WARN] Falha ao clicar no botão OK (tentativa {tentativa + 1})")
            continue

        # ── 4) Aguardar o diálogo ATUAL sumir após o clique ──
        try:
            WebDriverWait(driver, 5, poll_frequency=0.2).until(
                _dialog_desapareceu()
            )
        except TimeoutException:
            # Diálogo não sumiu — pode ser que o clique não funcionou
            # ou é um novo diálogo. Continua o loop.
            print("[WARN] Diálogo não sumiu após clique. Tentando novamente...")
            continue

        # Pequena pausa para dar tempo de um novo diálogo aparecer (se houver)
        time.sleep(0.5)

    print(f"[INFO] Finalizou com {cliques} clique(s) no botão OK.")
    return cliques


def _is_displayed_safe(element):
    """Verifica is_displayed() sem explodir em StaleElementReference."""
    try:
        return element.is_displayed()
    except (StaleElementReferenceException, Exception):
        return False


def _clicar_botao_ok(driver, btn):
    """
    Tenta clicar no botão usando 3 estratégias.
    Retorna o nome do método que funcionou, ou None se todos falharem.
    """
    # Estratégia 1: Clique direto
    try:
        btn.click()
        return "clique direto"
    except (ElementClickInterceptedException, ElementNotInteractableException) as e:
        print(f"  [fallback] Clique direto falhou: {type(e).__name__}")
    except StaleElementReferenceException:
        return None

    # Estratégia 2: ActionChains
    try:
        ActionChains(driver).move_to_element(btn).pause(0.1).click().perform()
        return "ActionChains"
    except Exception as e:
        print(f"  [fallback] ActionChains falhou: {type(e).__name__}")

    # Estratégia 3: JavaScript
    try:
        driver.execute_script("arguments[0].click();", btn)
        return "JavaScript"
    except Exception as e:
        print(f"  [fallback] JS click falhou: {type(e).__name__}")

    return None


class _dialog_desapareceu:
    """
    Condition customizada: retorna True quando NÃO há mais
    nenhum mat-dialog-container visível na página.
    """
    def __call__(self, driver):
        dialogs = driver.find_elements(By.CSS_SELECTOR, "mat-dialog-container")
        visiveis = [d for d in dialogs if _is_displayed_safe(d)]
        return len(visiveis) == 0
    
def wait_no_overlay(driver, wait):
    """Espera sumirem overlays/spinners do Angular Material."""
    try:
        wait.until(EC.invisibility_of_element_located((
            By.CSS_SELECTOR,
            ".cdk-overlay-backdrop.cdk-overlay-backdrop-showing, .mat-progress-spinner, .mat-progress-bar"
        )))
    except TimeoutException:
        pass

def js_set_value_and_dispatch(driver, el, value: str):
    """Define valor via JS e dispara eventos para Angular Reactive Forms perceber."""
    driver.execute_script("""
        const el = arguments[0], v = arguments[1];
        el.focus();
        const setter = Object.getOwnPropertyDescriptor(HTMLInputElement.prototype, 'value').set;
        setter.call(el, v);
        el.dispatchEvent(new Event('input', { bubbles: true }));
        el.dispatchEvent(new Event('change', { bubbles: true }));
        el.blur();
    """, el, value)

def get_input_by_formcontrol(driver, wait, formcontrol):
    sel = f"input[formcontrolname='{formcontrol}']"
    el = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, sel)))
    wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, sel)))
    scroll_center(driver, el)
    return el    

def esperar_transicao_apos_primeiro(wait, btn_clique=None, timeout=40):
    # Se temos referência do botão, esperar ele sumir ou ficar 'stale'
    if btn_clique is not None:
        try:
            wait.until(EC.staleness_of(btn_clique))
            return
        except TimeoutException:
            pass
        try:
            wait.until(EC.invisibility_of_element(btn_clique))
            return
        except TimeoutException:
            pass
    # fallback: pequena espera + overlays
    time.sleep(0.3)
    wait_overlays_to_hide(wait)

def click_with_fallback(driver, el):
    try:
        el.click()
    except (ElementClickInterceptedException, ElementNotInteractableException):
        driver.execute_script("arguments[0].click();", el)

def type_safely(driver, wait, el, value: str):
    """Tenta digitar e valida. Se não pegar, usa JS + eventos."""
    try:
        click_with_fallback(driver, el)
        el.send_keys(Keys.CONTROL, 'a', Keys.DELETE)
        if value is not None:
            el.send_keys(value)
        # valida
        time.sleep(0.05)
        v = el.get_attribute("value") or ""
        if v.strip() != (value or "").strip():
            js_set_value_and_dispatch(driver, el, value or "")
    except StaleElementReferenceException:
        # o chamador deve re-obter o elemento
        raise

def is_center_clickable_js(driver, el):
    """Verifica via JS se o centro do elemento está clicável (não coberto)."""
    try:
        return driver.execute_script("""
            const el = arguments[0];
            const r = el.getBoundingClientRect();
            if (r.width === 0 || r.height === 0) return false;
            const cx = r.left + r.width/2, cy = r.top + r.height/2;
            const e = document.elementFromPoint(cx, cy);
            return e && (e === el || el.contains(e));
        """, el)
    except Exception:
        return False

def debug_dump(driver, prefix):
    try: driver.save_screenshot(f"{prefix}.png")
    except: pass
    try:
        with open(f"{prefix}.html","w",encoding="utf-8") as f: f.write(driver.page_source)
    except: pass

def ensure_enabled_and_editable(driver, el, allow_force=False):
    """Se readonly/disabled, tenta remover o atributo e deixar editável (force=True para campos que vêm travados)."""
    readonly = el.get_attribute("readonly")
    disabled = el.get_attribute("disabled")
    if readonly is not None or disabled is not None:
        if not allow_force:
            # Apenas informa — pode ser campo calculado/auto-preenchido pelo sistema
            return False
        # Força habilitar visualmente e tenta setar via JS (nem sempre muda o estado do FormControl, mas costuma funcionar)
        driver.execute_script("arguments[0].removeAttribute('readonly'); arguments[0].removeAttribute('disabled');", el)
    return True

def click_primeiro_continuar(driver, wait, campo_desc_css='input[formcontrolname="frmDescNota"]'):
    # Garantir blur da descrição (habilita o botão se houver validação)
    try:
        campo_desc = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, campo_desc_css)))
        campo_desc.send_keys(Keys.TAB)
        time.sleep(0.1)
    except TimeoutException:
        print("[WARN] Campo de descrição não encontrado para TAB; seguindo.")

    wait_overlays_to_hide(wait)

    # Pega o primeiro <button> CONTINUAR visível e habilitado
    candidatos = driver.find_elements(By.XPATH, "//button[.//span[normalize-space(.)='CONTINUAR']]")
    if not candidatos:
        debug_dump(driver, "continuar_1_sem_candidatos")
        raise RuntimeError("Nenhum <button> CONTINUAR encontrado.")

    alvo = None
    for btn in candidatos:
        try:
            vis = btn.is_displayed()
            hab = btn.is_enabled() and btn.get_attribute("disabled") in (None, "", "false")
            aria = btn.get_attribute("aria-disabled")
            if vis and hab and (aria in (None, "", "false")) and "mat-button-disabled" not in (btn.get_attribute("class") or ""):
                alvo = btn
                break
        except Exception:
            continue

    if not alvo:
        # fallback: qualquer visível
        for btn in candidatos:
            try:
                if btn.is_displayed():
                    alvo = btn
                    break
            except Exception:
                continue

    if not alvo:
        debug_dump(driver, "continuar_1_sem_alvo")
        raise RuntimeError("Nenhum <button> CONTINUAR visível.")

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", alvo)
    time.sleep(0.05)

    try:
        wait.until(EC.element_to_be_clickable((By.XPATH, "//button[.//span[normalize-space(.)='CONTINUAR']]")))
    except TimeoutException:
        print("[INFO] element_to_be_clickable não confirmou; tentando mesmo assim.")

    try:
        if not is_center_clickable_js(driver, alvo):
            driver.execute_script("window.scrollBy(0, -80);")
            time.sleep(0.05)
        alvo.click()
        return alvo  # <<<< devolvemos o botão clicado para aguardar transição
    except Exception as e1:
        print(f"[INFO] Clique direto (btn) falhou: {repr(e1)}")

    try:
        ActionChains(driver).move_to_element(alvo).pause(0.05).click().perform()
        return alvo
    except Exception as e2:
        print(f"[INFO] ActionChains (btn) falhou: {repr(e2)}")

    try:
        driver.execute_script("arguments[0].click();", alvo)
        return alvo
    except Exception as e3:
        debug_dump(driver, "continuar_1_click_fail")
        raise RuntimeError(f"Falha ao clicar no primeiro CONTINUAR: {repr(e3)}")

def click_segundo_continuar(driver, wait):
    wait_overlays_to_hide(wait)

    locator_link = (By.XPATH,
        "//a[.//span[normalize-space(.)='CONTINUAR'] "
        " and (not(@aria-disabled) or @aria-disabled='false') "
        " and not(contains(@class,'mat-button-disabled'))]"
    )

    # Espera aparecer ao menos um <a> candidato
    try:
        wait.until(EC.presence_of_element_located(locator_link))
    except TimeoutException:
        debug_dump(driver, "continuar_2_sem_presenca")
        raise RuntimeError("Nenhum <a> CONTINUAR presente após o primeiro clique.")

    links = driver.find_elements(*locator_link)
    if not links:
        debug_dump(driver, "continuar_2_sem_links")
        raise RuntimeError("Nenhum <a> CONTINUAR encontrado (filtro).")

    # Escolhe o primeiro visível e habilitado
    alvo = None
    for a in links:
        try:
            vis = a.is_displayed()
            hab = a.is_enabled()
            aria = a.get_attribute("aria-disabled")
            cls = a.get_attribute("class") or ""
            if vis and hab and (aria in (None, "", "false")) and "mat-button-disabled" not in cls:
                alvo = a
                break
        except Exception:
            continue

    # fallback: qualquer <a> visível
    if not alvo:
        for a in links:
            try:
                if a.is_displayed():
                    alvo = a
                    break
            except Exception:
                continue

    if not alvo:
        debug_dump(driver, "continuar_2_sem_alvo")
        raise RuntimeError("Nenhum <a> CONTINUAR visível.")

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", alvo)
    time.sleep(0.05)

    try:
        wait.until(EC.element_to_be_clickable(locator_link))
    except TimeoutException:
        print("[INFO] Link CONTINUAR não confirmado clicável; tentando mesmo assim.")

    # Clique normal → ActionChains → JS
    try:
        if not is_center_clickable_js(driver, alvo):
            driver.execute_script("window.scrollBy(0, -80);")
            time.sleep(0.05)
        alvo.click()
        return
    except Exception as e1:
        print(f"[INFO] Clique direto (link) falhou: {repr(e1)}")

    try:
        ActionChains(driver).move_to_element(alvo).pause(0.05).click().perform()
        return
    except Exception as e2:
        print(f"[INFO] ActionChains (link) falhou: {repr(e2)}")

    try:
        driver.execute_script("arguments[0].click();", alvo)
        return
    except Exception as e3:
        debug_dump(driver, "continuar_2_click_fail")
        raise RuntimeError(f"Falha ao clicar no segundo CONTINUAR (<a>): {repr(e3)}")

def click_pesquisar(driver, wait):
    wait_overlays_to_hide(wait)

    locator_btn = (
        By.XPATH,
        "//button[.//span[normalize-space(.)='Pesquisar'] and not(@disabled) and not(contains(@class,'mat-button-disabled'))]"
    )

    try:
        btn = wait.until(EC.presence_of_element_located(locator_btn))
    except TimeoutException:
        debug_dump(driver, "pesquisar_nao_encontrado")
        raise RuntimeError("Botão 'Pesquisar' não encontrado na tela.")

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", btn)
    time.sleep(0.05)

    try:
        wait.until(EC.element_to_be_clickable(locator_btn))
    except TimeoutException:
        print("[INFO] 'Pesquisar' não confirmou clicável; tentando assim mesmo.")

    try:
        if not is_center_clickable_js(driver, btn):
            driver.execute_script("window.scrollBy(0, -80);")
            time.sleep(0.05)
        btn.click()
        return
    except Exception as e1:
        print(f"[INFO] Clique direto (Pesquisar) falhou: {repr(e1)}")

    try:
        ActionChains(driver).move_to_element(btn).pause(0.05).click().perform()
        return
    except Exception as e2:
        print(f"[INFO] ActionChains (Pesquisar) falhou: {repr(e2)}")

    try:
        driver.execute_script("arguments[0].click();", btn)
        return
    except Exception as e3:
        debug_dump(driver, "pesquisar_click_fail")
        raise RuntimeError(f"Falha ao clicar no botão 'Pesquisar': {repr(e3)}")

#------------------------------------------------------------------------------

def _aguardar_checkbox_interagivel(driver, timeout=40):
    """
    Aguarda até que o checkbox da primeira linha esteja:
      1. Presente no DOM
      2. Visível na tela
      3. Clicável (não coberto por overlay)
      4. Estável (não muda de posição — tabela terminou de renderizar)

    Retorna o elemento <input> do checkbox.
    """
    locator_checkbox = (
        By.XPATH,
        "(//td[contains(@class,'mat-column-select')]"
        "//mat-checkbox//input[@type='checkbox'])[1]"
    )

    # ── Passo 1: Esperar presença no DOM ──
    wait = WebDriverWait(
        driver, timeout,
        poll_frequency=0.3,
        ignored_exceptions=[StaleElementReferenceException],
    )

    try:
        cb_input = wait.until(EC.presence_of_element_located(locator_checkbox))
    except TimeoutException:
        raise RuntimeError(
            f"Checkbox não apareceu no DOM após {timeout}s."
        )

    # ── Passo 2: Esperar que o elemento esteja VISÍVEL ──
    try:
        cb_input = wait.until(EC.visibility_of_element_located(locator_checkbox))
    except TimeoutException:
        # Em Angular Material, o <input> real é "cdk-visually-hidden",
        # então a visibilidade do input em si pode falhar.
        # Nesse caso, esperamos a visibilidade do <mat-checkbox> pai.
        locator_mat_checkbox = (
            By.XPATH,
            "(//td[contains(@class,'mat-column-select')]//mat-checkbox)[1]"
        )
        try:
            wait.until(EC.visibility_of_element_located(locator_mat_checkbox))
        except TimeoutException:
            raise RuntimeError(
                f"mat-checkbox não ficou visível após {timeout}s."
            )

    # ── Passo 3: Esperar estabilidade de posição (tabela parou de renderizar) ──
    _aguardar_posicao_estavel(driver, cb_input, tentativas=5, intervalo=0.3)

    return cb_input


def _aguardar_posicao_estavel(driver, elemento, tentativas=5, intervalo=0.3):
    """
    Verifica se o elemento parou de se mover na tela.
    Isso garante que a tabela terminou de carregar/renderizar.
    """
    pos_anterior = None
    for _ in range(tentativas):
        try:
            rect = driver.execute_script(
                "var r = arguments[0].getBoundingClientRect();"
                "return {top: r.top, left: r.left, width: r.width, height: r.height};",
                elemento,
            )
        except StaleElementReferenceException:
            time.sleep(intervalo)
            continue

        if pos_anterior and rect == pos_anterior:
            return  # Posição estável
        pos_anterior = rect
        time.sleep(intervalo)


def _obter_label_do_checkbox(driver, cb_input):
    """
    Retorna o <label> associado ao checkbox, que é o alvo ideal de clique
    em Angular Material (o <input> é visually-hidden).
    """
    try:
        input_id = cb_input.get_attribute("id")
        if input_id:
            return driver.find_element(By.XPATH, f"//label[@for='{input_id}']")
    except Exception:
        pass

    # Fallback: subir até o <mat-checkbox> e pegar o <label> dentro dele
    try:
        return driver.find_element(
            By.XPATH,
            "(//td[contains(@class,'mat-column-select')]"
            "//mat-checkbox//label)[1]"
        )
    except Exception:
        return None


def _clicar_com_fallback(driver, alvo, descricao="elemento"):
    """
    Tenta clicar no elemento usando 3 estratégias em cascata:
      1. Clique direto do Selenium
      2. ActionChains
      3. JavaScript
    """
    # Estratégia 1: Clique direto
    try:
        alvo.click()
        return "Clique direto"
    except (ElementClickInterceptedException, Exception) as e:
        print(f"[INFO] Clique direto ({descricao}) falhou: {type(e).__name__}")

    # Estratégia 2: ActionChains
    try:
        ActionChains(driver).move_to_element(alvo).pause(0.1).click().perform()
        return "ActionChains"
    except Exception as e:
        print(f"[INFO] ActionChains ({descricao}) falhou: {type(e).__name__}")

    # Estratégia 3: JavaScript
    try:
        driver.execute_script("arguments[0].click();", alvo)
        return "Clique via JS"
    except Exception as e:
        raise RuntimeError(
            f"Todas as estratégias de clique falharam para {descricao}: {repr(e)}"
        )


def selecionar_primeira_linha_checkbox(
    driver,
    wait,
    timeout=40,
    textos_para_verificar=None,
    exigir_todos=False,
    clicar_mesmo_se_faltar=False,
):
    """
    Seleciona o checkbox da primeira linha da tabela Angular Material.
    """
    wait_overlays_to_hide(wait)

    # ── 1) Verificação opcional de textos ──
    resultado_textos = {
        "verificacao_feita": False,
        "encontrados": [],
        "faltando": [],
    }

    if textos_para_verificar:
        encontrados, faltando, _conteudo = verificar_textos_na_tabela(
            driver, wait, textos_para_verificar, timeout=timeout
        )
        resultado_textos.update({
            "verificacao_feita": True,
            "encontrados": encontrados,
            "faltando": faltando,
        })

        if faltando and exigir_todos and not clicar_mesmo_se_faltar:
            debug_dump(driver, "textos_nao_encontrados_na_tabela")
            raise RuntimeError(
                f"Textos faltando na tabela: {faltando} | Encontrados: {encontrados}"
            )

        if faltando and not clicar_mesmo_se_faltar:
            return {
                **resultado_textos,
                "clicou": False,
                "motivo": f"Faltando textos: {faltando}",
            }

    # ── 2) Aguardar checkbox REALMENTE interagível ──
    try:
        cb_input = _aguardar_checkbox_interagivel(driver, timeout=timeout)
    except RuntimeError:
        debug_dump(driver, "checkbox_nao_encontrado")
        raise

    # ── 3) Scroll até o checkbox ──
    driver.execute_script(
        "arguments[0].scrollIntoView({block:'center'});", cb_input
    )
    time.sleep(0.15)

    # ── 4) Obter o <label> (alvo preferencial de clique) ──
    label = _obter_label_do_checkbox(driver, cb_input)
    alvo = label if label else cb_input

    # ── 5) Verificar se não há overlay cobrindo ──
    try:
        if not is_center_clickable_js(driver, alvo):
            driver.execute_script("window.scrollBy(0, -100);")
            time.sleep(0.15)
    except Exception:
        pass  # Se a verificação falhar, tenta clicar mesmo assim

    # ── 6) Clicar com fallback ──
    metodo = _clicar_com_fallback(driver, alvo, descricao="checkbox primeira linha")

    return {
        **resultado_textos,
        "clicou": True,
        "motivo": metodo,
    }

def preencher_quantidade_e_valor(driver, wait, quantidade="1", valor_boleto="123,45"):
    wait_overlays_to_hide(wait)

    # QUANTIDADE
    locator_qtd = (
        By.XPATH,
        "//input[@id='quantidade' or @name='quantidade' or @formcontrolname='frmQuantidade']"
    )
    try:
        qtd = wait.until(EC.presence_of_element_located(locator_qtd))
    except TimeoutException:
        debug_dump(driver, "quantidade_nao_presente")
        raise RuntimeError("Campo 'Quantidade' não encontrado.")

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", qtd)
    time.sleep(0.05)

    try:
        wait.until(EC.element_to_be_clickable(locator_qtd))
    except TimeoutException:
        print("[INFO] Quantidade não confirmou clicável; tentando assim mesmo.")

    try:
        clear_and_type(qtd, quantidade)
    except Exception as e:
        print(f"[INFO] Falha ao digitar quantidade (tentando via JS): {repr(e)}")
        driver.execute_script("arguments[0].value = arguments[1]; arguments[0].dispatchEvent(new Event('input',{bubbles:true}));", qtd, quantidade)

    # VALOR UNITÁRIO (currency mask)
    locator_valor = (
        By.XPATH,
        "//input[@id='valorUnitario' or @name='valorUnitario' or @formcontrolname='frmValor']"
    )
    try:
        valor = wait.until(EC.presence_of_element_located(locator_valor))
    except TimeoutException:
        debug_dump(driver, "valor_nao_presente")
        raise RuntimeError("Campo 'Valor Unitário' não encontrado.")

    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", valor)
    time.sleep(0.05)

    try:
        wait.until(EC.element_to_be_clickable(locator_valor))
    except TimeoutException:
        print("[INFO] Valor unitário não confirmou clicável; tentando assim mesmo.")

    # Decida como enviar: só dígitos (máscara formata) ou pt-BR (xx,yy).
    # Aqui vou enviar como string do boleto; se a máscara não aceitar, tente apenas dígitos.
    try:
        clear_and_type(valor, valor_boleto)
    except Exception as e1:
        print(f"[INFO] Falha ao digitar valor_boleto (direto): {repr(e1)}. Tentando só dígitos.")
        somente_digitos = re.sub(r"\D", "", str(valor_boleto))
        try:
            clear_and_type(valor, somente_digitos)
        except Exception as e2:
            print(f"[INFO] Falha com dígitos (tentando via JS): {repr(e2)}")
            driver.execute_script(
                "arguments[0].value = arguments[1]; arguments[0].dispatchEvent(new Event('input',{bubbles:true}));",
                valor, somente_digitos
            )

def selecionar_opcao_justificativa_com_hover(driver, wait, texto_alvo="2 - Orientações do gestor/coordendor da área"):
    # Painel de opções deve estar aberto neste ponto
    # 1) Localiza a opção pelo texto (tolerante a espaços/acentos)
    locator_option_exata = (
        By.XPATH,
        f"//mat-option//span[contains(normalize-space(.), '{texto_alvo.split(' - ')[-1].split()[0]}') and contains(normalize-space(.), 'Orientações do gestor')]"
    )
    try:
        opt_span = wait.until(EC.presence_of_element_located(locator_option_exata))
    except TimeoutException:
        # Fallback: qualquer mat-option que contenha "Orientações do gestor"
        try:
            opt_span = wait.until(EC.presence_of_element_located((
                By.XPATH,
                "//mat-option//span[contains(normalize-space(.), 'Orientações do gestor')]"
            )))
        except TimeoutException:
            debug_dump(driver, "justificativa_opcao_nao_presente")
            raise RuntimeError("Não encontrei a opção de justificativa no painel.")

    # 2) Sobe do <span> para o <mat-option> (alvo do hover)
    try:
        mat_option = opt_span.find_element(By.XPATH, "./ancestor::mat-option[1]")
    except Exception:
        mat_option = opt_span

    # 3) Move o mouse (hover), pequena pausa e clica
    driver.execute_script("arguments[0].scrollIntoView({block:'center'});", mat_option)
    time.sleep(0.05)

    try:
        ActionChains(driver).move_to_element(mat_option).pause(0.15).click().perform()
        return
    except Exception as e1:
        print(f"[INFO] Hover+click na opção falhou: {repr(e1)}")

    # 4) Fallback: clique direto/JS
    try:
        mat_option.click()
        return
    except Exception:
        try:
            driver.execute_script("arguments[0].click();", mat_option)
            return
        except Exception as e2:
            print(f"[INFO] JS click na opção falhou: {repr(e2)}")

    # 5) Fallback final: por teclado (↓ até achar a opção, ENTER)
    try:
        from selenium.webdriver.common.keys import Keys
        # Foca o painel (ou o body) e envia algumas setas para baixo + Enter
        body = driver.find_element(By.TAG_NAME, "body")
        for _ in range(10):  # percorre algumas opções
            body.send_keys(Keys.ARROW_DOWN)
            time.sleep(0.05)
        body.send_keys(Keys.ENTER)
    except Exception as e3:
        debug_dump(driver, "justificativa_opcao_hover_falhou")
        raise RuntimeError(f"Falha ao selecionar a justificativa (mesmo com hover): {repr(e3)}")


#-------------------------------------------------------------------------------------------------------------------------------


# ==========================================
#              FUNÇÕES ÚTEIS
# ==========================================

def _strip_accents(s: str) -> str:
    if pd.isna(s) or s is None:
        return ''
    s = unicodedata.normalize('NFKD', str(s))
    return ''.join(ch for ch in s if not unicodedata.combining(ch))

def _norm_colname(s: str) -> str:
    return _strip_accents(str(s)).lower().strip()

def _clean_str(x) -> str:
    if pd.isna(x): return ''
    if isinstance(x, float):
        if x.is_integer(): return str(int(x))
        return str(x)
    s = str(x).strip()
    if s.endswith('.0'): s = s[:-2]
    return s

def _norm_coletor(x: str) -> str:
    up = _strip_accents(_clean_str(x)).upper()
    return re.sub(r'[^A-Z0-9]', '', up)

def _is_valid_coletor(coletor: str) -> bool:
    if pd.isna(coletor) or str(coletor).strip() == '' or str(coletor).strip().lower() == 'nan':
        return False
    c = _strip_accents(_clean_str(coletor)).upper()
    c_clean = re.sub(r'[^A-Z0-9]', '', c)
    if re.fullmatch(r'\d{6,}', c_clean): 
        return True
    if 6 <= len(c_clean) <= 12 and any(x.isalpha() for x in c_clean) and any(x.isdigit() for x in c_clean): 
        return True
    return False

def _tipo_de_coletor(coletor: str) -> str:
    c = _norm_coletor(coletor)
    if c == 'SEMCENTRODECUSTO': return '-'
    if re.fullmatch(r'\d+', c): return 'N'
    if any(x.isalpha() for x in c) and any(x.isdigit() for x in c): return 'K'
    return ''

def _clean_valor_series(s: pd.Series) -> pd.Series:
    def limpa_valor(val):
        if pd.isna(val) or str(val).strip() == '': return None
        if isinstance(val, (int, float)): return float(val)
        v = str(val).upper().replace('R$', '').replace('\xa0', '').replace(' ', '').strip()
        if '.' in v and ',' in v: v = v.replace('.', '').replace(',', '.')
        elif ',' in v: v = v.replace(',', '.')
        try: return float(v)
        except ValueError: return None
    return s.apply(limpa_valor)

# ==========================================
#     LEITURA: RATEIO RECEBIDO (ISOLAMENTO)
# ==========================================

def ler_rr_bruto(caminho_rr: Union[str, Path]) -> pd.DataFrame:
    xls = pd.ExcelFile(caminho_rr, engine='openpyxl')
    frames = []
    
    for sh in xls.sheet_names:
        df = pd.read_excel(xls, sheet_name=sh, header=None, engine='openpyxl')
        idx_ancora = -1
        max_knf = 0
        
        for j in range(df.shape[1]):
            col_str = df.iloc[:, j].astype(str).str.strip().str.upper()
            count_knf = col_str.isin(['K', 'N', 'F']).sum()
            if count_knf > max_knf and count_knf >= 2:
                max_knf = count_knf
                idx_ancora = j
                
        if idx_ancora == -1: continue
            
        df_regional = df.iloc[:, idx_ancora:].copy()
        df_regional.columns = range(df_regional.shape[1])
        
        col_coletor, col_subnum, col_valor = -1, -1, -1
        header_row_idx = -1
        
        for i, row in df_regional.head(20).iterrows():
            row_norm = [_norm_colname(str(x)) for x in row.values]
            if 'coletor' in row_norm or 'subnumero' in row_norm:
                for j in range(len(row_norm)):
                    val = row_norm[j]
                    if val == 'coletor' and col_coletor == -1: col_coletor = j
                    elif 'subnumero' in val and col_subnum == -1: col_subnum = j
                    elif 'valor' in val and 'servico' not in val and col_valor == -1: col_valor = j
                
                if col_valor == -1:
                    for j in range(len(row_norm)):
                        if 'valor' in row_norm[j] and col_valor == -1: col_valor = j
                        
                if col_coletor != -1 and col_valor != -1:
                    header_row_idx = i
                    break

        if header_row_idx != -1:
            tmp = df_regional.iloc[header_row_idx + 1:].copy()
            def safe_get(c): return tmp.iloc[:, c] if c != -1 and c < tmp.shape[1] else pd.Series(['']*len(tmp))
            tmp_clean = pd.DataFrame({
                'COLETOR_ORIG': safe_get(col_coletor),
                'SUBNUM_ORIG': safe_get(col_subnum),
                'VALOR': safe_get(col_valor)
            })
        else:
            col_coletor_tb = -1
            max_validos = 0
            for j in range(1, df_regional.shape[1]):
                mask = df_regional.iloc[:, j].astype(str).apply(_is_valid_coletor)
                qtd = mask.sum()
                if qtd > max_validos and qtd >= 2:
                    max_validos = qtd
                    col_coletor_tb = j
                    
            col_valor_tb = -1
            max_nums = 0
            for j in range(1, df_regional.shape[1]):
                if j == col_coletor_tb: continue
                nums = _clean_valor_series(df_regional.iloc[:, j]).notna().sum()
                if nums > max_nums:
                    max_nums = nums
                    col_valor_tb = j
                    
            if col_coletor_tb == -1 or col_valor_tb == -1: continue
                
            tmp_clean = pd.DataFrame({
                'COLETOR_ORIG': df_regional.iloc[:, col_coletor_tb],
                'SUBNUM_ORIG': pd.Series(['']*len(df_regional)), 
                'VALOR': df_regional.iloc[:, col_valor_tb]
            })

        tmp_clean['VALOR'] = _clean_valor_series(tmp_clean['VALOR'])
        tmp_clean = tmp_clean.dropna(subset=['VALOR'])
        
        def get_best_coletor(r):
            sub = _clean_str(r['SUBNUM_ORIG'])
            if _is_valid_coletor(sub): return sub
            col = _clean_str(r['COLETOR_ORIG'])
            if _is_valid_coletor(col): return col
            return None
            
        tmp_clean['COLETOR_FINAL'] = tmp_clean.apply(get_best_coletor, axis=1)
        tmp_clean = tmp_clean.dropna(subset=['COLETOR_FINAL'])
        
        if not tmp_clean.empty:
            tmp_clean['COLETOR'] = tmp_clean['COLETOR_FINAL'].apply(_norm_coletor)
            tmp_clean['TIPOCOLETOR'] = tmp_clean['COLETOR'].apply(_tipo_de_coletor)
            frames.append(tmp_clean[['TIPOCOLETOR', 'COLETOR', 'VALOR']])

    if frames:
        return pd.concat(frames, ignore_index=True)
    return pd.DataFrame(columns=['TIPOCOLETOR', 'COLETOR', 'VALOR'])

# ==========================================
#     LEITURA: CORREIOS (COLA-PEDAÇOS)
# ==========================================

def _extrair_coletor_de_titular(texto: str) -> str:
    if pd.isna(texto) or str(texto).strip() == '': return "SEM CENTRO DE CUSTO"
    t = _strip_accents(str(texto)).upper()
    
    m = re.search(r'(?<!\d)(\d{6,})(?!\d)', t)
    if m: return _norm_coletor(m.group(1))
    
    palavras = t.split()
    for p in palavras:
        p_clean = re.sub(r'[^A-Z0-9]', '', p)
        if 6 <= len(p_clean) <= 12 and any(c.isalpha() for c in p_clean) and any(c.isdigit() for c in p_clean):
            return p_clean
            
    for i in range(len(palavras) - 1):
        p1 = re.sub(r'[^A-Z0-9]', '', palavras[i])
        p2 = re.sub(r'[^A-Z0-9]', '', palavras[i+1])
        comb = p1 + p2
        if 8 <= len(comb) <= 12 and any(c.isalpha() for c in comb) and any(c.isdigit() for c in comb):
            return comb
            
    return "SEM CENTRO DE CUSTO"

def ler_correios_bruto(caminho_correios: Union[str, Path]) -> Tuple[pd.DataFrame, float]:
    df_raw = pd.read_excel(caminho_correios, header=None, engine='openpyxl')
    idx_header = -1
    col_titular, col_valor = -1, -1
    
    for i, row in df_raw.head(20).iterrows():
        row_norm = [_norm_colname(str(x)) for x in row.values]
        if any('titular do cartao' in c for c in row_norm) and any('valor do servico' in c for c in row_norm):
            idx_header = i
            for j, c in enumerate(row_norm):
                if 'titular do cartao' in c: col_titular = j
                if 'valor do servico' in c: col_valor = j
            break
            
    valor_liquido = 0.0
    idx_fim_tabela = len(df_raw)
    
    for i in range(len(df_raw) - 1, -1, -1):
        row_norm = [_norm_colname(str(x)) for x in df_raw.iloc[i].values]
        if any('valor liquido' in c for c in row_norm):
            idx_fim_tabela = i
            for j, c in enumerate(row_norm):
                if 'valor liquido' in c:
                    if i + 1 < len(df_raw):
                        val_raw = df_raw.iloc[i + 1, j]
                        valor_liquido = _clean_valor_series(pd.Series([val_raw])).iloc[0]
                    break
            break

    if idx_header == -1: 
        return pd.DataFrame(columns=['TIPOCOLETOR', 'COLETOR', 'VALOR']), valor_liquido
    
    df = df_raw.iloc[idx_header + 1 : idx_fim_tabela, [col_titular, col_valor]].copy()
    df.columns = ['TITULAR', 'VALOR']
    
    # IGNORA LINHAS DE ENCARGOS E DESCONTOS (Serão tratados matematicamente no final)
    mask_ignorar = df['TITULAR'].astype(str).str.upper().str.contains('ENCARGO|DESCONTO|CREDITO')
    df = df[~mask_ignorar].copy()
    
    df['VALOR'] = _clean_valor_series(df['VALOR'])
    df = df.dropna(subset=['VALOR'])
    df['COLETOR'] = df['TITULAR'].apply(_extrair_coletor_de_titular)
    df['TIPOCOLETOR'] = df['COLETOR'].apply(_tipo_de_coletor)
    
    return df[['TIPOCOLETOR', 'COLETOR', 'VALOR']], valor_liquido

# ==========================================
#     CONSTRUÇÃO E CONCILIAÇÃO FINAL
# ==========================================

def _formatar_planilha_final(arquivo_xlsx: Union[str, Path], sheet='Planilha1'):
    wb = load_workbook(arquivo_xlsx)
    if sheet not in wb.sheetnames:
        wb.close(); return
    ws = wb[sheet]
    header = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    try:
        idx_valor = header.index('VALOR') + 1
        idx_op = header.index('OPERACAO') + 1
    except ValueError:
        wb.close(); return
        
    for row in ws.iter_rows(min_row=2):
        cell_valor = row[idx_valor - 1]
        if isinstance(cell_valor.value, (int, float)):
            cell_valor.number_format = numbers.FORMAT_NUMBER_00
            
    wb.save(arquivo_xlsx)
    wb.close()

def gerar_rateio_pag(
    caminho_correios: Union[str, Path],
    caminho_rr: Union[str, Path],
    saida: Union[str, Path] = 'RATEIO PAG.xlsx',
    operacao_para_diagrama: int = 10,
    tolerancia_igual: float = 0.05, 
    debug: bool = True
) -> pd.DataFrame:

    df_rr_raw = ler_rr_bruto(caminho_rr)
    df_corr_raw, valor_liquido_correios = ler_correios_bruto(caminho_correios)

    # GARANTIA DE TIPO NUMÉRICO
    df_rr_raw['VALOR'] = pd.to_numeric(df_rr_raw['VALOR'], errors='coerce').fillna(0.0)
    df_corr_raw['VALOR'] = pd.to_numeric(df_corr_raw['VALOR'], errors='coerce').fillna(0.0)

    total_rr = float(df_rr_raw['VALOR'].sum()) if not df_rr_raw.empty else 0.0
    total_corr_soma = float(df_corr_raw['VALOR'].sum()) if not df_corr_raw.empty else 0.0
    total_corr = valor_liquido_correios if valor_liquido_correios > 0 else total_corr_soma

    if debug:
        print(f"[DEBUG] TOTAL RR               = R$ {total_rr:.2f}")
        print(f"[DEBUG] TOTAL CORREIOS (SOMA)  = R$ {total_corr_soma:.2f}")
        print(f"[DEBUG] TOTAL CORREIOS (LÍQ)   = R$ {valor_liquido_correios:.2f}")
        print(f"[DEBUG] DIFERENÇA (LÍQ - RR)   = R$ {round(total_corr - total_rr, 2):.2f}")

    # 1. BASE INICIAL: Tudo que a regional aprovou
    linhas_finais = df_rr_raw.to_dict('records') if not df_rr_raw.empty else []

    if abs(total_corr - total_rr) > tolerancia_igual:
        if debug: print("[DEBUG] Valores divergem. Buscando pacotes exatos faltantes...")
        
        rr_valores_exatos = df_rr_raw['VALOR'].round(2).tolist()
        
        # Saldo por Centro de Custo
        saldo_rr_cc = {}
        for _, row in df_rr_raw.iterrows():
            c = row['COLETOR']
            saldo_rr_cc[c] = saldo_rr_cc.get(c, 0.0) + row['VALOR']
            
        # A "Carteira" Global (Garante que nunca vamos adicionar pacotes a mais)
        saldo_global_rr = total_rr
        pacotes_faltantes = []
        
        for _, row in df_corr_raw.iterrows():
            c = row['COLETOR']
            v = round(row['VALOR'], 2)
            tipo = row['TIPOCOLETOR']
            
            matched = False
            
            # Fase 1: Tenta pagar usando o saldo do próprio Centro de Custo
            if c in saldo_rr_cc and saldo_rr_cc[c] >= v - 0.02:
                saldo_rr_cc[c] -= v
                saldo_global_rr -= v
                matched = True
                # Remove da lista de valores exatos para não usar duas vezes
                for i, rr_v in enumerate(rr_valores_exatos):
                    if abs(rr_v - v) <= 0.02:
                        rr_valores_exatos.pop(i)
                        break
            
            # Fase 2: Tenta pagar achando o valor exato (caso o CC tenha mudado de nome)
            if not matched:
                for i, rr_v in enumerate(rr_valores_exatos):
                    if abs(rr_v - v) <= 0.02:
                        rr_valores_exatos.pop(i)
                        saldo_global_rr -= v
                        matched = True
                        break
                        
            # Se não achou de jeito nenhum, vai para a fila de espera
            if not matched:
                pacotes_faltantes.append({
                    'TIPOCOLETOR': tipo,
                    'COLETOR': c,
                    'VALOR': v
                })
                
        # Fase 3: Absorção pela Carteira Global
        # Se a regional agrupou valores E mudou o nome, o pacote cai aqui.
        # Usamos o dinheiro que sobrou na carteira para pagar. Se a carteira zerar, aí sim é um pacote faltante!
        pacotes_adicionados = 0
        for pct in pacotes_faltantes:
            v = pct['VALOR']
            if saldo_global_rr >= v - 0.02:
                # A carteira cobre! (Falso positivo, a regional já pagou por isso)
                saldo_global_rr -= v
            else:
                # A carteira não cobre! Realmente faltou na planilha da regional.
                linhas_finais.append(pct)
                pacotes_adicionados += 1
                
        if debug: print(f"[DEBUG] Adicionados {pacotes_adicionados} pacotes exatos dos Correios.")

    # Transforma em DataFrame e agrupa para unificar linhas do mesmo CC
    final_base = pd.DataFrame(linhas_finais)
    if not final_base.empty:
        final_base = final_base.groupby(['TIPOCOLETOR', 'COLETOR'], as_index=False)['VALOR'].sum()

    # 3. RATEIO DE ENCARGOS E DESCONTOS (Matemática)
    soma_atual = final_base['VALOR'].sum() if not final_base.empty else 0.0
    diferenca_rateio = round(total_corr - soma_atual, 2)
    
    # Se sobrou diferença (Encargos) ou faltou (Descontos), rateia entre todos
    if abs(diferenca_rateio) > 0.02 and not final_base.empty:
        if debug: print(f"[DEBUG] Rateando R$ {diferenca_rateio:.2f} (Encargos/Descontos) proporcionalmente...")
        
        soma_validos = final_base['VALOR'].sum()
        if soma_validos > 0:
            # Calcula a proporção
            final_base['VALOR_ADD'] = (final_base['VALOR'] / soma_validos) * diferenca_rateio
            final_base['VALOR_ADD'] = final_base['VALOR_ADD'].round(2)
            
            # Ajuste de centavos (dízimas)
            diff_centavos = round(diferenca_rateio - final_base['VALOR_ADD'].sum(), 2)
            if diff_centavos != 0:
                idx_max = final_base['VALOR'].idxmax() 
                final_base.loc[idx_max, 'VALOR_ADD'] += diff_centavos 
                
            # Aplica o rateio
            final_base['VALOR'] += final_base['VALOR_ADD']
            final_base = final_base.drop(columns=['VALOR_ADD'])

    # Monta Tabela Final
    final = pd.DataFrame()
    if not final_base.empty:
        final['ITEM'] = [1] * len(final_base)
        final['TIPOCOLETOR'] = final_base['TIPOCOLETOR']
        final['COLETOR'] = final_base['COLETOR']
        final['OPERACAO'] = final_base['TIPOCOLETOR'].apply(lambda t: operacao_para_diagrama if t == 'N' else '')
        final['SUBNUMERO'] = ''
        final['VALOR'] = final_base['VALOR']
        final['DESCRICAO'] = ''

        # Ordenação (K -> N -> Faltantes no Final)
        final['__ord'] = final['TIPOCOLETOR'].map({'K': 0, 'N': 1}).fillna(2)
        final = final.sort_values(['__ord', 'COLETOR']).drop(columns='__ord').reset_index(drop=True)

    saida = Path(saida)
    with pd.ExcelWriter(saida, engine='openpyxl') as writer:
        final.to_excel(writer, sheet_name='Planilha1', index=False)

    _formatar_planilha_final(saida, 'Planilha1')

    if debug: print(f"[DEBUG] Arquivo gerado com sucesso: {saida.resolve()}")

    return final



# ==============================================================================
# 4. CLASSE PARA REDIRECIONAR OS PRINTS PARA A INTERFACE
# ==============================================================================
class PrintRedirector:
    def __init__(self, text_widget):
        self.text_widget = text_widget

    def write(self, text):
        self.text_widget.configure(state='normal')
        self.text_widget.insert(tk.END, text)
        self.text_widget.see(tk.END) # Rola a tela para baixo automaticamente
        self.text_widget.configure(state='disabled')

    def flush(self):
        pass


# ==============================================================================
# 5. INTERFACE GRÁFICA (GUI)
# ==============================================================================
class AppRoboMRV:
    def __init__(self, root):
        self.root = root
        self.root.title("Robô de Faturamento - Correios MRV")
        self.root.geometry("700x550")
        self.root.configure(padx=20, pady=20)

        # Estilo
        style = ttk.Style()
        style.configure("TButton", font=("Arial", 10, "bold"), padding=10)
        style.configure("TLabel", font=("Arial", 12))

        # Título
        lbl_titulo = ttk.Label(root, text="Automação de Pagamentos - Correios", font=("Arial", 16, "bold"))
        lbl_titulo.pack(pady=(0, 20))

        # Frame dos Botões
        frame_botoes = ttk.Frame(root)
        frame_botoes.pack(fill=tk.X, pady=10)

        # Botão 1
        self.btn_etapa1 = ttk.Button(frame_botoes, text="1. Gerar Rascunhos (Outlook)", command=self.iniciar_etapa_1)
        self.btn_etapa1.pack(side=tk.LEFT, expand=True, padx=5)

        # Botão 2
        self.btn_etapa2 = ttk.Button(frame_botoes, text="2. Gerar Planilha de Rateio", command=self.iniciar_etapa_2)
        self.btn_etapa2.pack(side=tk.LEFT, expand=True, padx=5)

        # Botão 3
        self.btn_etapa3 = ttk.Button(frame_botoes, text="3. Lançar Nota (Portal MRV)", command=self.iniciar_etapa_3)
        self.btn_etapa3.pack(side=tk.LEFT, expand=True, padx=5)

        # Label do Console
        lbl_console = ttk.Label(root, text="Console de Execução:", font=("Arial", 10, "bold"))
        lbl_console.pack(anchor=tk.W, pady=(20, 5))

        # Caixa de Texto (Console)
        self.console = tk.Text(root, height=15, bg="black", fg="white", font=("Consolas", 10))
        self.console.pack(fill=tk.BOTH, expand=True)
        self.console.configure(state='disabled')

        # Redirecionar o 'print' do Python para a caixa de texto
        sys.stdout = PrintRedirector(self.console)
        sys.stderr = PrintRedirector(self.console) # Captura erros também

        print(f"Olá, Pedro! Sistema iniciado e pronto para uso.")
        print("Selecione uma das etapas acima para começar.\n" + "-"*50)

    # --- Funções de Execução em Thread (Evita travar a tela) ---
    
    def iniciar_etapa_1(self):
        threading.Thread(target=self._rodar_etapa_1, daemon=True).start()

    def _rodar_etapa_1(self):
        self._desabilitar_botoes()
        print("\n>>> INICIANDO ETAPA 1: Rascunhos Outlook...")
        try:
            criar_rascunhos_correios()
        except Exception as e:
            print(f"[ERRO] Falha na Etapa 1: {e}")
        finally:
            self._habilitar_botoes()

    def iniciar_etapa_2(self):
        threading.Thread(target=self._rodar_etapa_2, daemon=True).start()

    def _rodar_etapa_2(self):
        self._desabilitar_botoes()
        print("\n>>> INICIANDO ETAPA 2: Geração de Rateio...")
        try:
            preparar_e_gerar_rateio()
        except Exception as e:
            print(f"[ERRO] Falha na Etapa 2: {e}")
        finally:
            self._habilitar_botoes()

    def iniciar_etapa_3(self):
        threading.Thread(target=self._rodar_etapa_3, daemon=True).start()

    def _rodar_etapa_3(self):
        self._desabilitar_botoes()
        print("\n>>> INICIANDO ETAPA 3: Lançamento no Portal MRV...")
        try:
            lancar_nota_fiscal()
        except Exception as e:
            print(f"[ERRO] Falha na Etapa 3: {e}")
        finally:
            self._habilitar_botoes()

    # --- Controles da Interface ---
    def _desabilitar_botoes(self):
        self.btn_etapa1.state(['disabled'])
        self.btn_etapa2.state(['disabled'])
        self.btn_etapa3.state(['disabled'])

    def _habilitar_botoes(self):
        self.btn_etapa1.state(['!disabled'])
        self.btn_etapa2.state(['!disabled'])
        self.btn_etapa3.state(['!disabled'])

# ==============================================================================
# 6. INICIALIZAÇÃO DO APLICATIVO
# ==============================================================================
if __name__ == "__main__":
    root = tk.Tk()
    app = AppRoboMRV(root)
    root.mainloop()
