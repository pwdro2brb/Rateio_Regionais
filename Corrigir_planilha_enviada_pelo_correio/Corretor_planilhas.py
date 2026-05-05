from openpyxl import load_workbook
import re
from pathlib import Path

# === Configuração: altere para o nome do seu arquivo ===
infile = Path('Corrigir_planilha_enviada_pelo_correio/2491343.xlsx')

wb = load_workbook(infile)

# Regex: pega valores como "R$ 1.234,56" e "-R$ 166,31"
re_brl = re.compile(r'^\s*([-−])?\s*R\$\s*([0-9\.,]+)\s*$', re.IGNORECASE)

for ws in wb.worksheets:
    for row in ws.iter_rows():
        for cell in row:
            v = cell.value
            if isinstance(v, str):
                m = re_brl.match(v)
                if m:
                    sign = -1 if m.group(1) else 1
                    num = m.group(2)
                    # Remove milhar quando há vírgula decimal (ex.: 1.234,56 -> 1234,56)
                    if ',' in num and '.' in num:
                        num = num.replace('.', '')
                    # Converte vírgula para ponto (ex.: 1234,56 -> 1234.56)
                    num = num.replace(',', '.')
                    try:
                        val = sign * float(num)
                        cell.value = val
                        cell.number_format = 'R$ #,##0.00'
                    except Exception:
                        pass  # Se não conseguir converter, mantém como está

outfile = infile.with_name(f"{infile.stem}_corrigido{infile.suffix}")
wb.save(outfile)
print(f"Arquivo salvo: {outfile}")